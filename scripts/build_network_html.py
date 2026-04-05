"""
build_network_html.py — generates self-contained ra_genre_network.html.

Reads:
  data/ra_full_export.xlsx  (Episodes + Genre Edges sheets)
  data/genre_musicology.json (musicological genre relationships)

Output:
  ra_genre_network.html     (single-file D3 visualization)
  genre_mapping_gaps.txt    (unmapped genres, if any)
"""

import json
import glob
import re
import html
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"


def _sanitize_html(html_str):
    """Strip script tags, ensure links open in new tab."""
    if not html_str:
        return None
    s = re.sub(r'<script[^>]*>.*?</script>', '', html_str, flags=re.DOTALL | re.IGNORECASE)
    s = re.sub(r'<a ', '<a target="_blank" rel="noopener" ', s)
    return s.strip() or None


def _split_content(html_str):
    """Split content into article + Q&A based on first <b>...?</b> tag."""
    if not html_str:
        return None, None
    # Find first bold tag containing a question mark
    m = re.search(r'<b>[^<]*\?[^<]*</b>', html_str)
    if m:
        article = html_str[:m.start()].strip()
        qa = html_str[m.start():].strip()
        # Clean up trailing whitespace / empty paragraphs from article
        article = re.sub(r'\s*(<br\s*/?>|\n)+\s*$', '', article)
        return article or None, qa or None
    return html_str, None


def _normalize_keywords(raw):
    """Normalize RA keywords: decode HTML entities, title case, deduplicate."""
    if not raw:
        return None
    seen = {}
    result = []
    for kw in raw.split(','):
        kw = html.unescape(kw).strip()
        if not kw:
            continue
        normalized = kw.title()
        key = normalized.lower()
        if key not in seen:
            seen[key] = True
            result.append(normalized)
    return ','.join(result) if result else None


def load_mixes():
    """Load mixes from raw JSON + Excel (genres, parsed tracks)."""
    # ── Raw JSON episodes ────────────────────────────────────────────────
    raw_map = {}
    for fpath in glob.glob(str(DATA_DIR / "raw" / "episode_*.json")):
        with open(fpath, encoding="utf-8") as f:
            d = json.load(f)
        pid = d.get("podcast_id") or d.get("id")
        if not pid:
            continue
        pid = str(pid)
        title = d.get("title") or ""
        mix_m = re.match(r'RA\.(\d+)', title)
        mix_number = mix_m.group(1) if mix_m else pid
        artist_obj = d.get("artist") or {}
        trans = d.get("translation") or {}
        raw_map[pid] = {
            "mix_number": mix_number,
            "artist": re.sub(r'^RA\.\d+\s*', '', title).strip()
                      or (artist_obj.get("name") if artist_obj.get("name") not in (None, "None") else None)
                      or "Unknown",
            "artist_id": artist_obj.get("id"),
            "date": (d.get("date") or "")[:10] or None,
            "duration": d.get("duration"),
            "blurb": (trans.get("blurb") or "").strip() or None,
            "article": None,  # set below
            "qa": None,       # set below
            "imageUrl": d.get("imageUrl"),
            "streamingUrl": d.get("streamingUrl") or None,
            "url": d.get("url"),
            "keywords": _normalize_keywords(d.get("keywords")),
        }
        sanitized = _sanitize_html(trans.get("content"))
        article, qa = _split_content(sanitized)
        raw_map[pid]["article"] = article
        raw_map[pid]["qa"] = qa

    # ── Genre edges from Excel ───────────────────────────────────────────
    wb = openpyxl.load_workbook(DATA_DIR / "ra_full_export.xlsx", read_only=True)

    ws_ge = wb["Genre Edges"]
    ge_rows = list(ws_ge.iter_rows(values_only=True))
    ge_headers = ge_rows[0]
    eid_col = ge_headers.index("entity_id")
    gc_col = ge_headers.index("genre_canonical")

    genre_map = {}
    for row in ge_rows[1:]:
        pid = str(row[eid_col])
        genre = row[gc_col]
        if pid not in genre_map:
            genre_map[pid] = []
        if genre not in genre_map[pid]:
            genre_map[pid].append(genre)

    # ── Parsed tracks from Excel ─────────────────────────────────────────
    ws_tr = wb["Tracks"]
    tr_rows = list(ws_tr.iter_rows(values_only=True))
    tr_headers = tr_rows[0]
    tc = {h: i for i, h in enumerate(tr_headers)}

    tracks_map = {}
    for row in tr_rows[1:]:
        pid = str(row[tc["podcast_id"]])
        if pid not in tracks_map:
            tracks_map[pid] = []
        t = {"artist": row[tc["artist"]], "title": row[tc["title"]]}
        if row[tc["label"]]:
            t["label"] = row[tc["label"]]
        tracks_map[pid].append(t)

    wb.close()

    # ── LLM cache (labels + genres) ─────────────────────────────────────
    llm_cache = {}
    # Prefer enriched cache with label_categories if available
    for llm_path in [
        DATA_DIR / "llm_genre_cache_with_categories.jsonl",
        DATA_DIR / "llm_genre_cache_normalized.jsonl",
    ]:
        if llm_path.exists():
            with open(llm_path, encoding="utf-8") as f:
                for line in f:
                    if line.strip():
                        obj = json.loads(line)
                        llm_cache[obj["podcast_id"]] = obj
            break

    # ── Merge ────────────────────────────────────────────────────────────
    mixes = []
    for pid in sorted(raw_map.keys(), key=lambda x: int(x), reverse=True):
        ep = raw_map[pid]
        ep["genres"] = genre_map.get(pid, [])
        ep["tracks"] = tracks_map.get(pid, [])
        ep["labels"] = llm_cache.get(pid, {}).get("labels", [])
        ep["label_categories"] = llm_cache.get(pid, {}).get("label_categories", {})
        ep["id"] = pid
        mixes.append(ep)

    return mixes


def check_gaps(mixes, hierarchy):
    """Find genres in mixes that aren't in the hierarchy."""
    node_ids = {n["id"] for n in hierarchy["nodes"]}
    all_mix_genres = set()
    for m in mixes:
        all_mix_genres.update(m["genres"])

    gaps = sorted(all_mix_genres - node_ids)
    if gaps:
        gap_path = ROOT / "genre_mapping_gaps.txt"
        with open(gap_path, "w") as f:
            f.write("Genres in Excel that have no node in genre_hierarchy.json:\n\n")
            for g in gaps:
                count = sum(1 for m in mixes if g in m["genres"])
                f.write(f"  {g} ({count} episodes)\n")
        print(f"WARNING: {len(gaps)} unmapped genres → genre_mapping_gaps.txt")
    else:
        print("All genres mapped successfully.")
    return gaps


def build_html(mixes, graph):
    # Recalculate node counts from actual mixes data
    genre_counts = {}
    for m in mixes:
        for g in m.get("genres", []):
            genre_counts[g] = genre_counts.get(g, 0) + 1
    for node in graph["nodes"]:
        node["count"] = genre_counts.get(node["id"], 0)

    nodes_json = json.dumps(graph["nodes"], ensure_ascii=False)
    # Map musicology 'strength' (1-5) to 'weight' for JS compatibility
    edges_for_js = []
    for e in graph["edges"]:
        ejs = {"source": e["source"], "target": e["target"], "weight": e.get("strength", e.get("weight", 1))}
        if "note" in e:
            ejs["note"] = e["note"]
        if "type" in e:
            ejs["type"] = e["type"]
        edges_for_js.append(ejs)
    edges_json = json.dumps(edges_for_js, ensure_ascii=False)
    mixes_json = json.dumps(mixes, ensure_ascii=False)

    total_mixes = len(mixes)
    total_genres = len(graph["nodes"])
    total_edges = len(graph["edges"])
    total_artists = len(set(m.get("artist", "") for m in mixes if m.get("artist")))

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>RA Podcast Mixes</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.9.0/d3.min.js"></script>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=Lexend:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
/* ═══════════════════════════════════════════════════════════════════
   歌舞伎町 2AM — Kabukicho After Midnight
   Neon reflections on wet asphalt. Dense light in narrow alleys.
   Information overload at Shinjuku station. The hum of vending machines.
   ═══════════════════════════════════════════════════════════════════ */

:root {{
  --night: #080a18;
  --night-mid: #0c0e1e;
  --night-surface: #101225;
  --night-elevated: #16182e;
  --glass: rgba(10, 12, 28, 0.82);
  --glass-border: rgba(255, 255, 255, 0.06);
  --glass-border-hover: rgba(255, 255, 255, 0.12);
  --neon-pink: #ff2d78;
  --neon-cyan: #00e5ff;
  --neon-amber: #ffaa00;
  --neon-violet: #7b61ff;
  --neon-green: #00ff88;
  --text-bright: #f0ecf8;
  --text-mid: #9590a8;
  --text-dim: #4e4968;
  --text-ghost: #2a2640;
  --font-display: 'Syne', sans-serif;
  --font-body: 'Lexend', sans-serif;
  --font-data: 'JetBrains Mono', monospace;
  --ease: cubic-bezier(0.22, 0.61, 0.36, 1);
}}

* {{ margin: 0; padding: 0; box-sizing: border-box; }}

body {{
  background: var(--night);
  color: var(--text-bright);
  font-family: var(--font-body);
  font-weight: 300;
  overflow: hidden;
  height: 100vh;
}}

/* Film grain overlay */
body::after {{
  content: '';
  position: fixed;
  inset: 0;
  opacity: 0.025;
  pointer-events: none;
  z-index: 9999;
  background-image: url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
  background-size: 128px 128px;
}}

/* Neon scrollbars */
::-webkit-scrollbar {{ width: 3px; }}
::-webkit-scrollbar-track {{ background: transparent; }}
::-webkit-scrollbar-thumb {{ background: rgba(255,45,120,0.25); border-radius: 3px; }}
::-webkit-scrollbar-thumb:hover {{ background: rgba(255,45,120,0.5); }}

/* ── Header ─────────────────────────────────────────────────── */
.header {{
  height: 56px;
  background: var(--glass);
  backdrop-filter: blur(20px);
  border-bottom: 1px solid transparent;
  border-image: linear-gradient(90deg, transparent 0%, var(--neon-pink) 30%, var(--neon-cyan) 70%, transparent 100%) 1;
  display: flex;
  align-items: center;
  padding: 0 28px;
  gap: 28px;
  position: relative;
  z-index: 10;
}}
.header .title {{
  font-family: var(--font-display);
  font-weight: 800;
  font-size: 16px;
  color: var(--text-bright);
  letter-spacing: 4px;
  text-shadow: 0 0 30px rgba(255,45,120,0.3);
}}
.header .title:hover {{
  text-shadow: 0 0 40px rgba(255,45,120,0.6), 0 0 80px rgba(0,229,255,0.2);
}}
.header .stat {{
  font-family: var(--font-data);
  font-size: 11px;
  color: var(--text-dim);
  letter-spacing: 0.5px;
}}
.header .stat span {{
  color: var(--neon-cyan);
  text-shadow: 0 0 10px rgba(0,229,255,0.3);
}}

/* ── Layout ─────────────────────────────────────────────────── */
.container {{
  display: flex;
  height: calc(100vh - 56px);
}}

/* ── Sidebar — glassmorphic ─────────────────────────────────── */
.sidebar {{
  width: 280px;
  min-width: 280px;
  background: var(--glass);
  backdrop-filter: blur(16px);
  border-right: 1px solid var(--glass-border);
  display: flex;
  flex-direction: column;
  overflow: hidden;
}}
.search-box {{
  position: relative;
  padding: 16px;
  border-bottom: 1px solid var(--glass-border);
}}
.search-clear {{
  position: absolute;
  right: 28px;
  top: 50%;
  transform: translateY(-50%);
  background: none;
  border: none;
  color: var(--text-dim);
  font-size: 12px;
  cursor: pointer;
  padding: 2px 4px;
  line-height: 1;
  transition: color 0.2s;
}}
.search-clear:hover {{ color: var(--text-bright); }}
.search-box input {{
  width: 100%;
  padding: 11px 36px 11px 16px;
  background: rgba(255,255,255,0.03);
  border: 1px solid var(--glass-border);
  border-radius: 12px;
  color: var(--text-bright);
  font-family: var(--font-body);
  font-size: 13px;
  font-weight: 300;
  outline: none;
  transition: all 0.3s var(--ease);
}}
.search-box input::placeholder {{ color: var(--text-dim); }}
.search-box input:focus {{
  border-color: var(--neon-cyan);
  box-shadow: 0 0 0 3px rgba(0,229,255,0.08), inset 0 0 20px rgba(0,229,255,0.03);
}}
.episode-list {{
  flex: 1;
  overflow-y: auto;
  padding: 4px 0;
}}

.ep-item {{
  padding: 9px 18px;
  cursor: pointer;
  font-size: 12px;
  font-weight: 300;
  color: var(--text-mid);
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
  transition: all 0.2s var(--ease);
  border-left: 2px solid transparent;
}}
.ep-item:hover {{
  background: rgba(255,45,120,0.04);
  color: var(--text-bright);
  box-shadow: inset 2px 0 0 var(--neon-pink);
  border-left-color: transparent;
}}
.ep-item.active {{
  background: rgba(255,45,120,0.08);
  color: var(--text-bright);
  border-left-color: var(--neon-pink);
  box-shadow: inset 2px 0 12px rgba(255,45,120,0.15);
}}
.ep-item .ep-num {{
  font-family: var(--font-data);
  color: var(--text-ghost);
  margin-right: 10px;
  font-size: 10px;
}}
.ep-item.active .ep-num {{
  color: var(--neon-pink);
  text-shadow: 0 0 8px rgba(255,45,120,0.5);
}}

/* ── Center panel ───────────────────────────────────────────── */
.center-panel {{
  flex: 1;
  display: flex;
  flex-direction: column;
  overflow: hidden;
  position: relative;
}}

/* Floating glassmorphic tab pills */
.center-tabs {{
  position: absolute;
  top: 14px;
  left: 50%;
  transform: translateX(-50%);
  z-index: 10;
  display: flex;
  gap: 2px;
  background: var(--glass);
  backdrop-filter: blur(16px);
  border-radius: 24px;
  border: 1px solid rgba(212,29,74,0.45);
  box-shadow: 0 0 12px 2px rgba(212,29,74,0.18), 0 0 32px 4px rgba(6,182,212,0.10);
  padding: 4px;
}}
.center-tab {{
  background: transparent;
  border: none;
  color: var(--text-dim);
  padding: 7px 22px;
  cursor: pointer;
  font-family: var(--font-display);
  font-size: 12px;
  font-weight: 500;
  letter-spacing: 0.8px;
  border-radius: 20px;
  transition: all 0.25s var(--ease);
}}
.center-tab:hover {{ color: var(--text-mid); }}
.center-tab.active {{
  color: var(--text-bright);
  background: rgba(255,45,120,0.15);
  box-shadow: 0 0 16px rgba(255,45,120,0.1);
}}

/* ── Graph area — ambient city glow ─────────────────────────── */
.graph-area {{
  flex: 1;
  position: relative;
  overflow: hidden;
  background:
    radial-gradient(ellipse at 30% 40%, rgba(123,97,255,0.04) 0%, transparent 50%),
    radial-gradient(ellipse at 70% 60%, rgba(0,229,255,0.03) 0%, transparent 50%),
    radial-gradient(ellipse at 50% 50%, rgba(255,45,120,0.02) 0%, transparent 40%),
    var(--night);
}}
.graph-area svg {{ width: 100%; height: 100%; }}

/* ── Graph search ──────────────────────────────────────────── */
.graph-search {{
  position: absolute;
  bottom: 20px;
  right: 20px;
  z-index: 20;
  width: min(280px, calc(100% - 40px));
}}
.graph-search-input {{
  width: 100%;
  padding: 10px 36px 10px 38px;
  background: var(--glass);
  backdrop-filter: blur(24px);
  border: 1px solid var(--glass-border);
  border-radius: 14px;
  color: var(--text-bright);
  font-family: var(--font-body);
  font-size: 13px;
  font-weight: 300;
  outline: none;
  transition: all 0.3s var(--ease);
}}
.graph-search-input::placeholder {{ color: var(--text-dim); }}
.graph-search-input:focus {{
  border-color: var(--neon-cyan);
  box-shadow: 0 0 0 3px rgba(0,229,255,0.08), 0 8px 32px rgba(0,0,0,0.3);
}}
.graph-search-icon {{
  position: absolute;
  left: 13px;
  top: 50%;
  transform: translateY(-50%);
  color: var(--text-dim);
  font-size: 13px;
  pointer-events: none;
  transition: color 0.2s;
}}
.graph-search:focus-within .graph-search-icon {{ color: var(--neon-cyan); }}
.graph-search-clear {{
  position: absolute;
  right: 12px;
  top: 50%;
  transform: translateY(-50%);
  background: none;
  border: none;
  color: var(--text-dim);
  font-size: 12px;
  cursor: pointer;
  padding: 2px 4px;
  line-height: 1;
  display: none;
  transition: color 0.2s;
}}
.graph-search-clear:hover {{ color: var(--text-bright); }}
.graph-search-dropdown {{
  position: absolute;
  bottom: calc(100% + 6px);
  left: 0;
  right: 0;
  background: var(--glass);
  backdrop-filter: blur(24px);
  border: 1px solid var(--glass-border);
  border-radius: 12px;
  max-height: 260px;
  overflow-y: auto;
  display: none;
  box-shadow: 0 12px 40px rgba(0,0,0,0.5);
}}
.graph-search-dropdown::-webkit-scrollbar {{ width: 4px; }}
.graph-search-dropdown::-webkit-scrollbar-thumb {{ background: var(--glass-border); border-radius: 2px; }}
.graph-search-item {{
  padding: 9px 16px;
  cursor: pointer;
  display: flex;
  align-items: center;
  gap: 10px;
  transition: background 0.15s;
  font-family: var(--font-body);
  font-size: 13px;
  color: var(--text-main);
}}
.graph-search-item:first-child {{ border-radius: 12px 12px 0 0; }}
.graph-search-item:last-child {{ border-radius: 0 0 12px 12px; }}
.graph-search-item:hover, .graph-search-item.active {{
  background: rgba(255,255,255,0.06);
  color: var(--text-bright);
}}
.graph-search-item .gs-dot {{
  width: 8px;
  height: 8px;
  border-radius: 50%;
  flex-shrink: 0;
}}
.graph-search-item .gs-count {{
  margin-left: auto;
  font-family: var(--font-data);
  font-size: 11px;
  color: var(--text-dim);
}}
.graph-search-item .gs-family {{
  font-size: 11px;
  color: var(--text-dim);
}}

/* ── Family district labels (in SVG) ───────────────────────── */
.family-label {{
  font-family: var(--font-display);
  font-weight: 700;
  text-anchor: middle;
  pointer-events: none;
  text-transform: uppercase;
  letter-spacing: 0.15em;
}}

/* ── Label filter area ──────────────────────────────────────── */
.label-filter-area {{
  flex: 1;
  padding: 60px 24px 24px;
  overflow-y: auto;
  font-family: var(--font-body);
  background: var(--night);
}}
.label-search-box {{
  position: relative;
  margin-bottom: 14px;
}}
.label-search-clear {{
  position: absolute;
  right: 12px;
  top: 50%;
  transform: translateY(-50%);
  background: none;
  border: none;
  color: var(--text-dim);
  font-size: 12px;
  cursor: pointer;
  padding: 2px 4px;
  line-height: 1;
  transition: color 0.2s;
}}
.label-search-clear:hover {{ color: var(--text-bright); }}
.label-search-box input {{
  width: 100%;
  padding: 10px 36px 10px 16px;
  background: rgba(255,255,255,0.03);
  border: 1px solid var(--glass-border);
  border-radius: 12px;
  color: var(--text-bright);
  font-family: var(--font-body);
  font-size: 13px;
  font-weight: 300;
  outline: none;
  transition: all 0.3s var(--ease);
}}
.label-search-box input::placeholder {{ color: var(--text-dim); }}
.label-search-box input:focus {{
  border-color: var(--neon-cyan);
  box-shadow: 0 0 0 3px rgba(0,229,255,0.08);
}}

/* Active filters */
.active-filters {{
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 6px;
  min-height: 32px;
  margin-bottom: 18px;
  padding-bottom: 14px;
  border-bottom: 1px solid var(--glass-border);
}}
.active-filters:empty {{ display: none; }}
.active-filter-chip.label-chip {{
  font-size: 12px;
  padding: 5px 10px 5px 12px;
  font-style: normal;
  border-radius: 20px;
  cursor: pointer;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  transition: all 0.2s var(--ease);
}}
.active-filter-chip:hover {{ opacity: 0.8; }}
.active-filter-chip .remove {{ opacity: 0.4; font-size: 14px; }}
.active-filter-chip .remove:hover {{ opacity: 1; }}
.clear-filters-btn {{
  font-size: 11px;
  color: var(--text-dim);
  cursor: pointer;
  padding: 4px 14px;
  border: 1px solid var(--glass-border);
  border-radius: 20px;
  background: none;
  font-family: var(--font-body);
  transition: all 0.2s var(--ease);
}}
.clear-filters-btn:hover {{
  color: var(--neon-pink);
  border-color: var(--neon-pink);
  box-shadow: 0 0 12px rgba(255,45,120,0.15);
}}
.filter-match-count {{
  font-family: var(--font-data);
  font-size: 11px;
  color: var(--text-dim);
  margin-left: auto;
}}

/* Genre filter section */
.genre-filter-section {{
  margin-bottom: 20px;
  padding: 16px 18px;
  background: var(--night-surface);
  border: 1px solid var(--glass-border);
  border-radius: 14px;
  position: relative;
  overflow: hidden;
}}
.genre-filter-section::before {{
  content: '';
  position: absolute;
  top: 0;
  left: 0;
  right: 0;
  height: 1px;
  background: linear-gradient(90deg, transparent, var(--neon-pink), var(--neon-violet), transparent);
  opacity: 0.4;
}}
.genre-filter-title {{
  font-family: var(--font-display);
  font-size: 10px;
  text-transform: uppercase;
  color: var(--text-dim);
  margin-bottom: 12px;
  letter-spacing: 3px;
  font-weight: 600;
}}
.genre-filter-items {{
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
}}
.genre-filter-chip {{
  font-size: 12px;
  padding: 5px 12px;
  border-radius: 20px;
  cursor: pointer;
  color: #fff;
  opacity: 0.6;
  font-family: var(--font-body);
  font-weight: 400;
  border: none;
  transition: all 0.25s var(--ease);
}}
.genre-filter-chip:hover {{
  opacity: 0.95;
  box-shadow: 0 0 14px currentColor;
}}
.genre-filter-chip.selected {{
  opacity: 1;
  box-shadow: 0 0 0 1.5px rgba(255,255,255,0.5), 0 0 20px currentColor;
}}
.genre-filter-chip .chip-count {{
  font-family: var(--font-data);
  font-size: 9px;
  opacity: 0.5;
  margin-left: 4px;
}}

/* Masonry */
.label-masonry {{
  display: flex;
  gap: 12px;
}}

.label-column {{
  flex: 1;
  min-width: 0;
  display: flex;
  flex-direction: column;
}}

/* Group separator line */
.label-cat-group-sep {{
  height: 1px;
  background: linear-gradient(90deg, transparent, rgba(255,255,255,0.07), transparent);
  margin: 0;
  border: none;
}}

/* Category tile — top neon glow line */
.label-cat-section {{
  margin-bottom: 16px;
  padding: 14px 16px 10px;
  background: var(--night-surface);
  border: 1px solid var(--glass-border);
  border-radius: 14px;
  position: relative;
  overflow: hidden;
  flex-shrink: 0;
}}
.label-cat-section::before {{
  content: '';
  position: absolute;
  top: 0;
  left: 0;
  right: 0;
  height: 2px;
}}
.label-cat-section[data-cat="mood"]::before,
.label-cat-section[data-cat="energy"]::before,
.label-cat-section[data-cat="vibe"]::before {{ height: 3px; }}
.label-cat-section[data-cat="mood"]::before      {{ background: linear-gradient(90deg, transparent, #b87ee6, transparent); }}
.label-cat-section[data-cat="energy"]::before    {{ background: linear-gradient(90deg, transparent, #ff9f43, transparent); }}
.label-cat-section[data-cat="setting"]::before   {{ background: linear-gradient(90deg, transparent, #45b7d1, transparent); }}
.label-cat-section[data-cat="geography"]::before {{ background: linear-gradient(90deg, transparent, #26de81, transparent); }}
.label-cat-section[data-cat="style"]::before     {{ background: linear-gradient(90deg, transparent, #6bb5ff, transparent); }}
.label-cat-section[data-cat="era"]::before       {{ background: linear-gradient(90deg, transparent, #ffd93d, transparent); }}
.label-cat-section[data-cat="vibe"]::before      {{ background: linear-gradient(90deg, transparent, #ff6b9d, transparent); }}
.label-cat-section[data-cat="__kw__"]::before   {{ background: linear-gradient(90deg, transparent, #888888, transparent); }}

/* Glow shadow — primary group (mood, energy, vibe) */
.label-cat-section[data-cat="mood"]      {{ box-shadow: 0 -2px 20px rgba(184,126,230,0.18); }}
.label-cat-section[data-cat="energy"]    {{ box-shadow: 0 -2px 20px rgba(255,159,67,0.18); }}
.label-cat-section[data-cat="vibe"]      {{ box-shadow: 0 -2px 20px rgba(255,107,157,0.18); }}
/* Context group (geography, setting, style) */
.label-cat-section[data-cat="setting"]   {{ box-shadow: 0 -2px 16px rgba(69,183,209,0.08); }}
.label-cat-section[data-cat="geography"] {{ box-shadow: 0 -2px 16px rgba(38,222,129,0.08); }}
.label-cat-section[data-cat="style"]     {{ box-shadow: 0 -2px 16px rgba(107,181,255,0.08); }}
/* Metadata group (era, __kw__) */
.label-cat-section[data-cat="era"]       {{ box-shadow: 0 -2px 16px rgba(255,217,61,0.08); }}
.label-cat-section[data-cat="__kw__"]    {{ box-shadow: 0 -2px 16px rgba(136,136,136,0.08); }}

.label-cat-title {{
  font-family: var(--font-display);
  font-size: 10px;
  text-transform: uppercase;
  color: var(--text-dim);
  margin-bottom: 10px;
  letter-spacing: 3px;
  font-weight: 600;
}}
/* Metadata group opacity (era, __kw__) */
.label-cat-section[data-cat="era"],
.label-cat-section[data-cat="__kw__"] {{
  opacity: 0.75;
  transition: opacity 0.2s var(--ease);
}}
.label-cat-section[data-cat="era"]:hover,
.label-cat-section[data-cat="__kw__"]:hover {{
  opacity: 1;
}}

.label-cat-items {{
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
}}
.label-filter-area .filter-label-chip.label-chip {{
  font-size: 12px;
  padding: 4px 11px;
  border-radius: 20px;
  cursor: pointer;
  border: 1px solid;
  opacity: 0.55;
  font-style: normal;
  font-weight: 400;
  transition: all 0.25s var(--ease);
}}
.label-filter-area .filter-label-chip:hover {{
  opacity: 0.9;
}}
.label-filter-area .filter-label-chip.selected {{
  opacity: 1;
  box-shadow: 0 0 0 1px rgba(255,255,255,0.4), 0 0 12px currentColor;
}}
.label-filter-area .filter-label-chip .chip-count {{
  font-family: var(--font-data);
  font-size: 9px;
  opacity: 0.4;
  margin-left: 3px;
}}
.show-more-btn {{
  font-size: 11px;
  color: var(--text-ghost);
  cursor: pointer;
  padding: 4px 14px;
  background: none;
  border: 1px solid var(--glass-border);
  border-radius: 20px;
  font-family: var(--font-body);
  margin-top: 8px;
  transition: all 0.2s var(--ease);
}}
.show-more-btn:hover {{
  color: var(--text-mid);
  border-color: var(--glass-border-hover);
}}

/* Active filter chip for genres */
.active-filter-chip.genre-chip {{
  font-size: 12px;
  padding: 5px 10px 5px 12px;
  border-radius: 20px;
  cursor: pointer;
  color: #fff;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  border: none;
}}

/* ── Edge filter buttons ────────────────────────────────────── */
.edge-filters {{
  position: absolute;
  bottom: 16px;
  left: 50%;
  transform: translateX(-50%);
  display: flex;
  gap: 4px;
  background: var(--glass);
  backdrop-filter: blur(16px);
  padding: 4px;
  border-radius: 14px;
  border: 1px solid var(--glass-border);
}}
.edge-btn {{
  padding: 5px 14px;
  font-size: 11px;
  font-family: var(--font-body);
  background: transparent;
  border: none;
  border-radius: 10px;
  color: var(--text-dim);
  cursor: pointer;
  transition: all 0.2s var(--ease);
}}
.edge-btn:hover {{ color: var(--text-bright); }}
.edge-btn.active {{
  color: var(--neon-cyan);
  background: rgba(0,229,255,0.1);
  box-shadow: 0 0 10px rgba(0,229,255,0.1);
}}

/* ── Tooltip ────────────────────────────────────────────────── */
.tooltip {{
  position: fixed;
  pointer-events: none;
  background: var(--glass);
  backdrop-filter: blur(20px);
  border: 1px solid var(--glass-border-hover);
  border-radius: 10px;
  padding: 8px 16px;
  font-family: var(--font-body);
  font-size: 12px;
  color: var(--text-bright);
  white-space: nowrap;
  display: none;
  box-shadow: 0 8px 32px rgba(0,0,0,0.5), 0 0 1px rgba(255,45,120,0.3);
  z-index: 9999;
  max-width: 320px;
}}
.tooltip.has-desc {{
  white-space: normal;
  line-height: 1.4;
}}
.tooltip .tt-title {{
  font-weight: 600;
  margin-bottom: 4px;
}}
.tooltip .tt-desc {{
  opacity: 0.8;
  font-size: 11px;
}}

/* ── Node labels ────────────────────────────────────────────── */
.node-label {{
  fill: var(--text-bright);
  text-anchor: middle;
  pointer-events: none;
  font-family: var(--font-display);
  font-weight: 500;
  paint-order: stroke;
  stroke: var(--night);
  stroke-width: 3px;
  stroke-linejoin: round;
}}

/* ── Pulse animation ────────────────────────────────────────── */
@keyframes pulse-ring {{
  0% {{ stroke-opacity: 0.7; stroke-width: 3; }}
  100% {{ stroke-opacity: 0; stroke-width: 12; }}
}}
@keyframes idle-pulse {{
  0%, 100% {{ opacity: 0.08; }}
  50% {{ opacity: 0.25; }}
}}
.idle-pulse {{
  animation: idle-pulse 4s ease-in-out infinite;
}}
.pulse {{ animation: pulse-ring 1.5s ease-out infinite; }}

/* ── Right panel — glassmorphic ─────────────────────────────── */
.detail-panel {{
  width: 380px;
  min-width: 380px;
  background: var(--glass);
  backdrop-filter: blur(20px);
  border-left: 1px solid var(--glass-border);
  display: flex;
  flex-direction: column;
  overflow-y: auto;
  overflow-x: hidden;
  transition: width 0.35s var(--ease), min-width 0.35s var(--ease);
}}
.detail-panel.collapsed {{
  width: 0;
  min-width: 0;
  border-left: none;
}}
.detail-header {{
  padding: 20px 22px 14px;
  border-bottom: 1px solid var(--glass-border);
}}
.detail-header .ep-title {{
  font-family: var(--font-display);
  font-size: 20px;
  font-weight: 700;
  color: var(--text-bright);
  margin-bottom: 4px;
  letter-spacing: 0.5px;
}}
.detail-header .ep-artist {{
  font-size: 14px;
  font-weight: 300;
  color: var(--text-mid);
  margin-bottom: 10px;
}}
.detail-meta {{
  display: flex;
  gap: 16px;
  font-family: var(--font-data);
  font-size: 10px;
  color: var(--text-dim);
  margin-bottom: 12px;
}}
.detail-meta .meta-item {{ display: flex; align-items: center; gap: 4px; }}
.detail-meta .meta-val {{ color: var(--text-mid); }}

/* Genre chips */
.genre-chips {{
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  margin-bottom: 6px;
}}
.genre-chip {{
  font-size: 11px;
  padding: 3px 10px;
  border-radius: 20px;
  color: #fff;
  opacity: 0.85;
  cursor: pointer;
  transition: all 0.2s var(--ease);
}}
.genre-chip:hover {{ opacity: 1; box-shadow: 0 0 12px currentColor; }}

/* Cover image — neon bleed gradient */
.detail-cover {{
  width: 100%;
  height: 200px;
  overflow: hidden;
  border-bottom: 1px solid var(--glass-border);
  position: relative;
}}
.detail-cover::after {{
  content: '';
  position: absolute;
  bottom: 0; left: 0; right: 0;
  height: 80px;
  background:
    linear-gradient(180deg, transparent 0%, rgba(10,12,28,0.82) 100%),
    linear-gradient(0deg, rgba(255,45,120,0.06) 0%, transparent 40%);
  pointer-events: none;
}}
.detail-cover img {{
  width: 100%;
  height: 100%;
  object-fit: cover;
  object-position: top;
}}

/* Links */
.detail-links {{
  display: flex;
  gap: 8px;
  margin-bottom: 10px;
}}
.detail-link {{
  font-size: 11px;
  padding: 4px 14px;
  border-radius: 20px;
  background: rgba(255,255,255,0.04);
  border: 1px solid var(--glass-border);
  color: var(--text-mid);
  text-decoration: none;
  font-family: var(--font-body);
  transition: all 0.2s var(--ease);
}}
.detail-link:hover {{
  color: var(--neon-cyan);
  border-color: rgba(0,229,255,0.3);
  box-shadow: 0 0 16px rgba(0,229,255,0.1);
}}

/* Keyword chips */
.kw-chips {{
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  margin-top: 8px;
}}
.kw-chip {{
  font-size: 10px;
  padding: 2px 9px;
  border-radius: 20px;
  border: 1px solid var(--glass-border);
  color: var(--text-dim);
}}

/* Label chips in detail */
.label-chips {{
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  margin-top: 8px;
}}
.label-chip {{
  font-size: 10px;
  padding: 2px 9px;
  border-radius: 20px;
  border: 1px solid var(--glass-border);
  color: var(--text-mid);
  background: rgba(255,255,255,0.03);
  cursor: pointer;
  transition: all 0.2s var(--ease);
}}
.label-chip:hover {{ opacity: 0.8; }}
/* Category neon colors */
.label-chip[data-cat="mood"]      {{ background: rgba(184,126,230,0.08); border-color: rgba(184,126,230,0.25); color: #caa0f4; }}
.label-chip[data-cat="energy"]    {{ background: rgba(255,159,67,0.08);  border-color: rgba(255,159,67,0.25);  color: #ffbe78; }}
.label-chip[data-cat="setting"]   {{ background: rgba(69,183,209,0.08);  border-color: rgba(69,183,209,0.25);  color: #78d0e5; }}
.label-chip[data-cat="geography"] {{ background: rgba(38,222,129,0.08);  border-color: rgba(38,222,129,0.25);  color: #5aeca0; }}
.label-chip[data-cat="style"]     {{ background: rgba(107,181,255,0.08); border-color: rgba(107,181,255,0.25); color: #98c8ff; }}
.label-chip[data-cat="era"]       {{ background: rgba(255,217,61,0.08);  border-color: rgba(255,217,61,0.25);  color: #ffe580; }}
.label-chip[data-cat="vibe"]      {{ background: rgba(255,107,157,0.08); border-color: rgba(255,107,157,0.25); color: #ff98bc; }}
.label-chip[data-cat="__kw__"]    {{ background: rgba(136,136,136,0.08); border-color: rgba(136,136,136,0.25); color: #b8b8b8; }}
.label-cat-group {{ display: contents; }}
.label-cat-dot {{
  display: inline-block;
  width: 6px; height: 6px;
  border-radius: 50%;
  margin-right: 3px;
  vertical-align: middle;
}}

/* ── Blurb ──────────────────────────────────────────────────── */
.detail-blurb {{
  padding: 14px 22px;
  font-size: 12px;
  line-height: 1.7;
  color: var(--text-mid);
  font-style: italic;
  font-weight: 300;
  border-bottom: 1px solid var(--glass-border);
}}

/* ── Article ────────────────────────────────────────────────── */
.detail-article {{
  padding: 16px 22px;
  font-size: 12px;
  line-height: 1.8;
  color: var(--text-mid);
  font-weight: 300;
  border-bottom: 1px solid var(--glass-border);
}}
.detail-article b {{ color: var(--text-bright); font-weight: 500; }}
.detail-article i {{ color: var(--text-mid); }}
.detail-article a {{ color: var(--neon-cyan); text-decoration: none; }}
.detail-article a:hover {{ text-shadow: 0 0 8px rgba(0,229,255,0.4); }}
.detail-section-header {{
  padding: 16px 22px 6px;
  font-family: var(--font-display);
  font-size: 10px;
  font-weight: 600;
  color: var(--text-ghost);
  text-transform: uppercase;
  letter-spacing: 3px;
  border-top: 1px solid var(--glass-border);
}}
.detail-qa b {{
  display: block;
  margin-top: 14px;
  margin-bottom: 4px;
  color: var(--text-bright);
  font-size: 12px;
  font-weight: 500;
}}

/* ── Tracklist ──────────────────────────────────────────────── */
.tracklist-header {{
  padding: 14px 22px 8px;
  font-family: var(--font-display);
  font-size: 10px;
  font-weight: 600;
  color: var(--text-ghost);
  text-transform: uppercase;
  letter-spacing: 3px;
  position: sticky;
  top: 0;
  background: var(--glass);
  backdrop-filter: blur(16px);
  z-index: 1;
}}
.track-item {{
  padding: 8px 22px;
  font-size: 12px;
  line-height: 1.5;
  font-weight: 300;
  border-bottom: 1px solid rgba(255,255,255,0.03);
  transition: background 0.2s var(--ease);
}}
.track-item:hover {{ background: rgba(255,45,120,0.03); }}
.track-item .track-num {{
  font-family: var(--font-data);
  color: var(--text-ghost);
  display: inline-block;
  width: 28px;
  text-align: right;
  margin-right: 12px;
  font-size: 10px;
}}
.track-item .track-artist {{ color: var(--text-bright); font-weight: 400; }}
.track-item .track-title {{ color: var(--text-mid); }}
.track-item .track-label {{ color: var(--text-dim); font-style: italic; }}
.no-tracklist {{
  padding: 22px;
  color: var(--text-ghost);
  font-size: 12px;
  font-style: italic;
}}

/* ── Detail placeholder ─────────────────────────────────────── */
.detail-placeholder {{
  display: flex;
  align-items: center;
  justify-content: center;
  height: 100%;
  color: var(--text-ghost);
  font-family: var(--font-display);
  font-size: 13px;
  letter-spacing: 2px;
  text-align: center;
  padding: 24px;
}}

/* ── Close button ───────────────────────────────────────────── */
.detail-close {{
  position: absolute;
  top: 16px;
  right: 16px;
  background: rgba(255,255,255,0.04);
  border: 1px solid var(--glass-border);
  border-radius: 8px;
  color: var(--text-dim);
  font-size: 11px;
  padding: 4px 12px;
  cursor: pointer;
  font-family: var(--font-body);
  transition: all 0.2s var(--ease);
}}
.detail-close:hover {{
  color: var(--neon-pink);
  border-color: rgba(255,45,120,0.3);
  box-shadow: 0 0 12px rgba(255,45,120,0.1);
}}

/* ── Sidebar player ─────────────────────────────────────────── */
.sidebar-player {{
  border-top: 1px solid var(--glass-border);
  background: var(--night-surface);
  padding: 0;
  min-height: 0;
}}
.sidebar-player .player-info {{
  padding: 8px 18px 4px;
  font-size: 10px;
  cursor: pointer;
  transition: color 0.2s;
}}
.sidebar-player .player-info:hover .player-artist {{
  color: var(--neon-cyan);
}}
.sidebar-player .player-ep {{ color: var(--text-ghost); }}
.sidebar-player .player-artist {{ color: var(--text-bright); font-weight: 500; }}
.sidebar-player iframe {{
  width: 100%;
  border: none;
  display: block;
}}
.sidebar-player.sc iframe {{ height: 125px; }}
.sidebar-player.mc iframe {{ height: 120px; }}
.sidebar-player:empty {{ display: none; }}

/* ══════════════════════════════════════════════════════════════════
   MOBILE — max-width 768px
   ══════════════════════════════════════════════════════════════════ */
@media (max-width: 768px) {{
  /* Hide desktop-only elements */
  .sidebar {{ display: none; }}
  .center-tabs {{ display: none; }}

  .container {{
    flex-direction: column;
    height: calc(100vh - 56px);
    height: calc(100dvh - 56px);
    padding-bottom: 56px; /* space for fixed tab bar */
    padding-bottom: calc(56px + max(0px, env(safe-area-inset-bottom)));
  }}

  .center-panel {{
    width: 100%;
    flex: 1;
    min-height: 0; /* allow flex shrink */
  }}

  /* ── Graph area — full available space ──────────────────────── */
  #graphArea {{
    opacity: 1.0;
  }}
  #graph {{
    pointer-events: auto;
  }}

  /* Hide tooltip on mobile */
  #tooltip {{
    display: none !important;
  }}

  /* Node tap feedback */
  #graph circle:active {{
    filter: brightness(1.3) drop-shadow(0 0 8px rgba(255, 255, 255, 0.5));
  }}

  /* ── Graph search — repositioned for mobile ────────────────── */
  .graph-search {{
    bottom: auto;
    top: 12px;
    right: 12px;
    left: 12px;
    width: auto;
    z-index: 20;
  }}
  .graph-search.hidden-by-panel {{
    opacity: 0;
    pointer-events: none;
  }}

  /* ── Detail panel → bottom sheet ───────────────────────────── */
  .detail-panel {{
    position: fixed;
    bottom: 56px;
    left: 0;
    right: 0;
    width: 100% !important;
    min-width: 0 !important;
    height: 50vh;
    border-left: none;
    border-top: 1px solid rgba(255, 45, 120, 0.3);
    border-radius: 20px 20px 0 0;
    transform: translateY(100%);
    transition: transform 0.35s var(--ease);
    z-index: 200;
    overflow: hidden;
    display: flex;
    flex-direction: column;
  }}
  .detail-panel.expanded {{
    height: 90vh;
  }}
  #detailContent {{
    overflow-y: auto;
    overflow-x: hidden;
    flex: 1;
  }}
  .detail-panel.collapsed {{
    transform: translateY(100%);
    width: 100% !important;
  }}
  .detail-panel:not(.collapsed) {{
    transform: translateY(0);
  }}

  /* Drag handle */
  .sheet-handle {{
    width: 36px;
    height: 4px;
    background: rgba(255, 255, 255, 0.2);
    border-radius: 2px;
    margin: 12px auto 8px;
    flex-shrink: 0;
    position: sticky;
    top: 0;
    z-index: 10;
  }}

  /* ── Mobile bottom tab bar ─────────────────────────────────── */
  .mobile-tab-bar {{
    display: flex;
    position: fixed;
    bottom: 0;
    left: 0;
    right: 0;
    height: 56px;
    padding-bottom: max(0px, env(safe-area-inset-bottom));
    background: rgba(8, 10, 24, 0.95);
    backdrop-filter: blur(20px);
    border-top: 1px solid rgba(212, 29, 74, 0.3);
    z-index: 300;
  }}
  .mobile-tab {{
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 3px;
    color: rgba(255, 255, 255, 0.4);
    font-family: var(--font-display);
    font-size: 10px;
    background: none;
    border: none;
    cursor: pointer;
    transition: color 0.2s;
    padding: 0;
  }}
  .mobile-tab.active {{
    color: var(--neon-pink);
  }}
  .mobile-tab svg {{
    width: 20px;
    height: 20px;
    stroke: currentColor;
    fill: none;
    stroke-width: 1.5;
  }}

  /* Header — safe area + hide stats */
  .header {{
    padding: max(0px, env(safe-area-inset-top)) 16px 0 16px;
  }}
  .header .stat {{
    display: none;
  }}

  /* ── Explore — single column ───────────────────────────────── */
  .label-masonry {{
    flex-direction: column;
  }}

  /* ── Mixes Tab ─────────────────────────────────────────────── */
  #mixesArea {{
    display: none;
    flex: 1;
    flex-direction: column;
    overflow: hidden;
  }}
  #mixesArea.active-view {{
    display: flex;
  }}
  #mixesList {{
    flex: 1;
    overflow-y: auto;
    padding: 8px 0;
  }}
  .mix-row {{
    display: flex;
    align-items: center;
    padding: 12px 18px;
    cursor: pointer;
    border-bottom: 1px solid var(--glass-border);
    transition: background 0.15s;
    gap: 0;
  }}
  .mix-row:active {{
    background: rgba(255, 45, 120, 0.08);
  }}
  .mix-row-num {{
    font-family: var(--font-data);
    font-size: 10px;
    color: var(--neon-pink);
    min-width: 56px;
    flex-shrink: 0;
  }}
  .mix-row-artist {{
    flex: 1;
    font-size: 13px;
    color: var(--text-bright);
    font-weight: 400;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
    margin-right: 8px;
  }}
  .mix-row-date {{
    font-family: var(--font-data);
    font-size: 10px;
    color: var(--text-dim);
    white-space: nowrap;
    flex-shrink: 0;
  }}
  #mixesDetail {{
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow-y: auto;
    overflow-x: hidden;
  }}
  .mixes-detail-header {{
    display: flex;
    align-items: center;
    padding: 12px 16px;
    border-bottom: 1px solid var(--glass-border);
    background: var(--glass);
    backdrop-filter: blur(16px);
    flex-shrink: 0;
    position: sticky;
    top: 0;
    z-index: 5;
  }}
  .mixes-back-btn {{
    background: none;
    border: none;
    color: var(--neon-pink);
    font-family: var(--font-display);
    font-size: 13px;
    cursor: pointer;
    padding: 4px 0;
    display: flex;
    align-items: center;
    gap: 6px;
    letter-spacing: 0.5px;
  }}
  .mixes-detail-content {{
    flex: 1;
    overflow: visible;
  }}
}}

/* Desktop: hide mobile elements */
@media (min-width: 769px) {{
  .mobile-tab-bar {{
    display: none;
  }}
  .sheet-handle {{
    display: none;
  }}
  #mixesArea {{
    display: none !important;
  }}
}}
</style>
</head>
<body>

<div class="header">
  <span class="title" style="cursor:pointer" onclick="resetApp()">RA Podcast Mixes</span>
  <span class="stat"><span>{total_mixes}</span> mixes</span>
  <span class="stat"><span>{total_artists}</span> artists</span>
  <span class="stat"><span>{total_genres}</span> genres</span>
</div>

<div class="container">
  <!-- Left: episode list -->
  <div class="sidebar">
    <div class="search-box">
      <input type="text" id="search" placeholder="Search episodes...">
      <button class="search-clear" id="searchClear" style="display:none">✕</button>
    </div>
    <div class="episode-list" id="episodeList"></div>
    <div class="sidebar-player" id="sidebarPlayer"></div>
  </div>

  <!-- Center: tabs + views -->
  <div class="center-panel">
    <div class="center-tabs">
      <button class="center-tab active" data-view="graph">Genre Map</button>
      <button class="center-tab" data-view="labels">Explore</button>
    </div>
    <div class="graph-area" id="graphArea">
      <svg id="graph"></svg>
      <div class="graph-search" id="graphSearch">
        <span class="graph-search-icon">⌕</span>
        <input type="text" class="graph-search-input" id="graphSearchInput" placeholder="Search genres..." autocomplete="off">
        <button class="graph-search-clear" id="graphSearchClear">✕</button>
        <div class="graph-search-dropdown" id="graphSearchDropdown"></div>
      </div>
    </div>
    <div class="tooltip" id="tooltip"></div>
    <div class="label-filter-area" id="labelFilterArea" style="display:none">
      <div class="label-search-box">
        <input type="text" id="labelSearch" placeholder="Search labels and genres...">
        <button class="label-search-clear" id="labelSearchClear" style="display:none">✕</button>
      </div>
      <div class="active-filters" id="activeFilters"></div>
      <div class="genre-filter-section" id="genreFilterSection"></div>
      <div class="label-masonry" id="labelMasonry"></div>
    </div>
    <div id="mixesArea">
      <div id="mixesList"></div>
      <div id="mixesDetail" style="display:none"></div>
    </div>
  </div>

  <!-- Right: episode detail -->
  <div class="detail-panel collapsed" id="detailPanel">
    <div class="sheet-handle"></div>
    <div id="detailContent">
      <div class="detail-placeholder">Select an episode to see details</div>
    </div>
  </div>
</div>

<!-- Mobile bottom tab bar -->
<nav class="mobile-tab-bar" id="mobileTabBar">
  <button class="mobile-tab" data-view="mixes">
    <svg viewBox="0 0 24 24">
      <line x1="4" y1="6" x2="20" y2="6"/>
      <line x1="4" y1="12" x2="20" y2="12"/>
      <line x1="4" y1="18" x2="20" y2="18"/>
    </svg>
    <span>Mixes</span>
  </button>
  <button class="mobile-tab active" data-view="graph">
    <svg viewBox="0 0 24 24">
      <circle cx="6" cy="6" r="1.5"/><circle cx="18" cy="6" r="1.5"/><circle cx="12" cy="12" r="1.5"/>
      <circle cx="6" cy="18" r="1.5"/><circle cx="18" cy="18" r="1.5"/>
      <line x1="6" y1="6" x2="12" y2="12"/><line x1="18" y1="6" x2="12" y2="12"/>
      <line x1="6" y1="18" x2="12" y2="12"/><line x1="18" y1="18" x2="12" y2="12"/>
    </svg>
    <span>Genre Map</span>
  </button>
  <button class="mobile-tab" data-view="labels">
    <svg viewBox="0 0 24 24">
      <circle cx="12" cy="12" r="9"/>
      <circle cx="12" cy="12" r="5"/>
      <circle cx="12" cy="12" r="1"/>
    </svg>
    <span>Explore</span>
  </button>
</nav>

<script>
// ── Embedded Data ───────────────────────────────────────────────────────────
const GRAPH_NODES = {nodes_json};
const GRAPH_EDGES = {edges_json};
const MIXES = {mixes_json};

// ── Lookups ─────────────────────────────────────────────────────────────────
const nodeMap = new Map(GRAPH_NODES.map(n => [n.id, n]));
const mixMap = new Map(MIXES.map(m => [m.id, m]));

// Adjacency: genre → set of neighbor genres
const adjacency = new Map();
GRAPH_EDGES.forEach(e => {{
  if (!adjacency.has(e.source)) adjacency.set(e.source, new Set());
  if (!adjacency.has(e.target)) adjacency.set(e.target, new Set());
  adjacency.get(e.source).add(e.target);
  adjacency.get(e.target).add(e.source);
}});

const maxWeight = Math.max(...GRAPH_EDGES.map(e => e.weight));
const maxCount = Math.max(...GRAPH_NODES.map(n => n.count));
function nodeRadius(count) {{ return 4 + 18 * Math.sqrt(count / maxCount); }}


// ── Episode List ────────────────────────────────────────────────────────────
const episodeList = document.getElementById('episodeList');
const searchInput = document.getElementById('search');
const detailPanel = document.getElementById('detailPanel');
const detailContent = document.getElementById('detailContent');

function renderEpisodeList(filter = '') {{
  const lower = filter.toLowerCase();
  episodeList.innerHTML = '';
  MIXES.forEach(m => {{
    if (filter && !m.artist.toLowerCase().includes(lower) &&
        !m.id.includes(filter)) return;
    const div = document.createElement('div');
    div.className = 'ep-item';
    div.dataset.id = m.id;
    div.innerHTML = `<span class="ep-num">RA.${{m.mix_number.padStart(3,'0')}}</span>${{m.artist}}`;
    div.addEventListener('click', () => selectEpisode(m));
    episodeList.appendChild(div);
  }});
}}

function renderEpisodeListByGenres(genreSet) {{
  episodeList.innerHTML = '';
  MIXES.forEach(m => {{
    if (!m.genres.some(g => genreSet.has(g))) return;
    const div = document.createElement('div');
    div.className = 'ep-item';
    div.dataset.id = m.id;
    div.innerHTML = `<span class="ep-num">RA.${{m.mix_number.padStart(3,'0')}}</span>${{m.artist}}`;
    div.addEventListener('click', () => selectEpisode(m));
    episodeList.appendChild(div);
  }});
}}

const searchClear = document.getElementById('searchClear');
searchInput.addEventListener('input', e => {{
  searchClear.style.display = e.target.value ? 'block' : 'none';
  renderEpisodeList(e.target.value);
}});
if (searchClear) {{
  searchClear.addEventListener('click', () => {{
    searchInput.value = '';
    searchClear.style.display = 'none';
    renderEpisodeList('');
    searchInput.focus();
  }});
}}
renderEpisodeList();

// ── Label & Genre Filter Panel ─────────────────────────────────────────────
const CAT_ORDER = ['mood','setting','era','energy','geography','vibe','style'];
const CAT_NAMES = {{ mood:'Mood', energy:'Energy', vibe:'Vibe', style:'Style', geography:'Geography', setting:'Setting', era:'Era' }};
const CAT_GROUPS = [
  ['mood', 'energy', 'vibe'],
  ['geography', 'setting', 'style'],
  ['era'],
];
const DEFAULT_SHOW = 80;

// Precompute label counts per category
const LABEL_COUNTS = {{}};
CAT_ORDER.forEach(cat => {{ LABEL_COUNTS[cat] = {{}}; }});
MIXES.forEach(m => {{
  if (!m.label_categories) return;
  CAT_ORDER.forEach(cat => {{
    (m.label_categories[cat] || []).forEach(l => {{
      LABEL_COUNTS[cat][l] = (LABEL_COUNTS[cat][l] || 0) + 1;
    }});
  }});
}});

// Precompute RA keyword counts
const KW_COUNTS = {{}};
MIXES.forEach(m => {{
  if (!m.keywords) return;
  m.keywords.split(',').map(k => k.trim()).filter(Boolean).forEach(k => {{
    KW_COUNTS[k] = (KW_COUNTS[k] || 0) + 1;
  }});
}});

// Precompute genre counts
const GENRE_COUNTS = {{}};
MIXES.forEach(m => {{
  (m.genres || []).forEach(g => {{
    GENRE_COUNTS[g] = (GENRE_COUNTS[g] || 0) + 1;
  }});
}});
const SORTED_GENRES = Object.entries(GENRE_COUNTS).sort((a, b) => b[1] - a[1]);

let activeFilters = [];  // [{{category, label}}]  category='genre' for genres
let expandedCats = new Set();

function getFilteredMixes() {{
  if (activeFilters.length === 0) return MIXES;
  return MIXES.filter(m => {{
    const cats = m.label_categories || {{}};
    return activeFilters.every(f => {{
      if (f.category === 'genre') return (m.genres || []).includes(f.label);
      if (f.category === '__kw__') return m.keywords ? m.keywords.split(',').map(k => k.trim()).includes(f.label) : false;
      return (cats[f.category] || []).includes(f.label);
    }});
  }});
}}

function computeFilteredCounts(filteredMixes) {{
  const labelCounts = {{}};
  CAT_ORDER.forEach(cat => {{ labelCounts[cat] = {{}}; }});
  const genreCounts = {{}};
  const kwCounts = {{}};
  filteredMixes.forEach(m => {{
    if (m.label_categories) {{
      CAT_ORDER.forEach(cat => {{
        (m.label_categories[cat] || []).forEach(l => {{
          labelCounts[cat][l] = (labelCounts[cat][l] || 0) + 1;
        }});
      }});
    }}
    (m.genres || []).forEach(g => {{
      genreCounts[g] = (genreCounts[g] || 0) + 1;
    }});
    if (m.keywords) {{
      m.keywords.split(',').map(k => k.trim()).filter(Boolean).forEach(k => {{
        kwCounts[k] = (kwCounts[k] || 0) + 1;
      }});
    }}
  }});
  return {{ labelCounts, genreCounts, kwCounts }};
}}

function chipFontSize(count) {{
  return Math.round(9 + Math.log2(Math.max(count, 1)) * 1.1);
}}

function renderGenreFilterSection() {{
  const container = document.getElementById('genreFilterSection');
  container.innerHTML = '';

  const filteredMixes = getFilteredMixes();
  const displayCounts = activeFilters.length > 0
    ? computeFilteredCounts(filteredMixes).genreCounts
    : GENRE_COUNTS;

  let entries = Object.entries(displayCounts).sort((a, b) => b[1] - a[1]);
  if (labelSearchQuery) entries = entries.filter(([g]) => g.toLowerCase().includes(labelSearchQuery));
  if (!entries.length) return;

  const expanded = expandedCats.has('genre');
  const visible = expanded ? entries : entries.slice(0, 30);
  const hasMore = entries.length > 30;

  const title = document.createElement('div');
  title.className = 'genre-filter-title';
  title.textContent = `Genres (${{entries.length}})`;
  container.appendChild(title);

  const items = document.createElement('div');
  items.className = 'genre-filter-items';

  visible.forEach(([genre, count]) => {{
    const chip = document.createElement('span');
    chip.className = 'genre-filter-chip';
    const node = nodeMap.get(genre);
    const color = node ? node.color : '#666';
    chip.style.background = color;
    chip.style.fontSize = chipFontSize(count) + 'px';
    chip.innerHTML = `${{genre}}<span class="chip-count">${{count}}</span>`;
    if (activeFilters.some(f => f.category === 'genre' && f.label === genre)) {{
      chip.classList.add('selected');
    }}
    chip.addEventListener('click', () => toggleLabelFilter('genre', genre));
    const desc = node && node.description ? node.description : '';
    if (desc) {{
      chip.addEventListener('mouseenter', (e) => {{
        const tt = document.getElementById('tooltip');
        tt.innerHTML = `<div class="tt-title">${{genre}}</div><div class="tt-desc">${{desc}}</div>`;
        tt.classList.add('has-desc');
        tt.style.display = 'block';
        requestAnimationFrame(() => positionTooltip(e));
      }});
      chip.addEventListener('mousemove', (e) => positionTooltip(e));
      chip.addEventListener('mouseleave', () => {{
        const tt = document.getElementById('tooltip');
        tt.style.display = 'none';
        tt.classList.remove('has-desc');
      }});
    }}
    items.appendChild(chip);
  }});

  container.appendChild(items);

  if (hasMore) {{
    const btn = document.createElement('button');
    btn.className = 'show-more-btn';
    btn.textContent = expanded ? 'Show less' : `Show all ${{entries.length}}`;
    btn.addEventListener('click', () => {{
      if (expanded) expandedCats.delete('genre');
      else expandedCats.add('genre');
      renderFilterPanel();
    }});
    container.appendChild(btn);
  }}
}}

function renderLabelMasonry() {{
  const container = document.getElementById('labelMasonry');
  container.innerHTML = '';

  // Create 3 columns
  const col1 = document.createElement('div');
  col1.className = 'label-column';
  const col2 = document.createElement('div');
  col2.className = 'label-column';
  const col3 = document.createElement('div');
  col3.className = 'label-column';

  // Column mapping: mood, setting, era, __kw__ → col1; energy, geography → col2; vibe, style → col3
  const columnMap = {{
    mood: col1, setting: col1, era: col1,
    energy: col2, geography: col2,
    vibe: col3, style: col3
  }};

  const filteredMixes = getFilteredMixes();
  const displayCounts = activeFilters.length > 0
    ? computeFilteredCounts(filteredMixes).labelCounts
    : LABEL_COUNTS;

  // Render categories into their assigned columns
  CAT_ORDER.forEach(cat => {{
      let entries = Object.entries(displayCounts[cat] || {{}})
        .sort((a, b) => b[1] - a[1]);
      if (labelSearchQuery) entries = entries.filter(([l]) => l.toLowerCase().includes(labelSearchQuery));
      if (!entries.length) return;

      const expanded = expandedCats.has(cat);
      const visible = expanded ? entries : entries.slice(0, DEFAULT_SHOW);
      const hasMore = entries.length > DEFAULT_SHOW;

      const section = document.createElement('div');
      section.className = 'label-cat-section';
      section.setAttribute('data-cat', cat);

      const title = document.createElement('div');
      title.className = 'label-cat-title';
      title.textContent = `${{CAT_NAMES[cat]}} (${{entries.length}})`;
      section.appendChild(title);

      const items = document.createElement('div');
      items.className = 'label-cat-items';

      visible.forEach(([label, count]) => {{
        const chip = document.createElement('span');
        chip.className = 'filter-label-chip label-chip';
        chip.setAttribute('data-cat', cat);
        chip.style.fontSize = chipFontSize(count) + 'px';
        chip.innerHTML = `${{label}}<span class="chip-count">${{count}}</span>`;
        if (activeFilters.some(f => f.category === cat && f.label === label)) {{
          chip.classList.add('selected');
        }}
        chip.addEventListener('click', () => toggleLabelFilter(cat, label));
        items.appendChild(chip);
      }});

      section.appendChild(items);

      if (hasMore) {{
        const btn = document.createElement('button');
        btn.className = 'show-more-btn';
        btn.textContent = expanded ? 'Show less' : `Show all ${{entries.length}}`;
        btn.addEventListener('click', () => {{
          if (expanded) expandedCats.delete(cat);
          else expandedCats.add(cat);
          renderFilterPanel();
        }});
        section.appendChild(btn);
      }}

      columnMap[cat].appendChild(section);
  }});

  // RA Tags section → col1
  const displayKwCounts = activeFilters.length > 0
    ? computeFilteredCounts(filteredMixes).kwCounts
    : KW_COUNTS;
  let kwEntries = Object.entries(displayKwCounts).sort((a, b) => b[1] - a[1]);
  if (labelSearchQuery) kwEntries = kwEntries.filter(([k]) => k.toLowerCase().includes(labelSearchQuery));
  if (kwEntries.length > 0) {{
    const expanded = expandedCats.has('__kw__');
    const visible = expanded ? kwEntries : kwEntries.slice(0, DEFAULT_SHOW);
    const hasMore = kwEntries.length > DEFAULT_SHOW;

    const section = document.createElement('div');
    section.className = 'label-cat-section';
    section.setAttribute('data-cat', '__kw__');

    const title = document.createElement('div');
    title.className = 'label-cat-title';
    title.textContent = `RA Tags (${{kwEntries.length}})`;
    section.appendChild(title);

    const items = document.createElement('div');
    items.className = 'label-cat-items';

    visible.forEach(([kw, count]) => {{
      const chip = document.createElement('span');
      chip.className = 'filter-label-chip label-chip';
      chip.setAttribute('data-cat', '__kw__');
      chip.style.fontSize = chipFontSize(count) + 'px';
      chip.innerHTML = `${{kw}}<span class="chip-count">${{count}}</span>`;
      if (activeFilters.some(f => f.category === '__kw__' && f.label === kw)) {{
        chip.classList.add('selected');
      }}
      chip.addEventListener('click', () => toggleLabelFilter('__kw__', kw));
      items.appendChild(chip);
    }});

    section.appendChild(items);

    if (hasMore) {{
      const btn = document.createElement('button');
      btn.className = 'show-more-btn';
      btn.textContent = expanded ? 'Show less' : `Show all ${{kwEntries.length}}`;
      btn.addEventListener('click', () => {{
        if (expanded) expandedCats.delete('__kw__');
        else expandedCats.add('__kw__');
        renderFilterPanel();
      }});
      section.appendChild(btn);
    }}

    col1.appendChild(section);
  }}

  container.appendChild(col1);
  container.appendChild(col2);
  container.appendChild(col3);
}}

let labelSearchQuery = '';

function renderFilterPanel() {{
  renderGenreFilterSection();
  renderLabelMasonry();
}}

// Label search input
const labelSearchInput = document.getElementById('labelSearch');
const labelSearchClear = document.getElementById('labelSearchClear');
if (labelSearchInput) {{
  labelSearchInput.addEventListener('input', (e) => {{
    labelSearchQuery = e.target.value.toLowerCase();
    labelSearchClear.style.display = labelSearchQuery ? 'block' : 'none';
    renderFilterPanel();
  }});
}}
if (labelSearchClear) {{
  labelSearchClear.addEventListener('click', () => {{
    labelSearchInput.value = '';
    labelSearchQuery = '';
    labelSearchClear.style.display = 'none';
    renderFilterPanel();
    labelSearchInput.focus();
  }});
}}

function renderActiveFilters() {{
  const container = document.getElementById('activeFilters');
  container.innerHTML = '';
  if (activeFilters.length === 0) return;

  activeFilters.forEach((f, idx) => {{
    const chip = document.createElement('span');
    if (f.category === 'genre') {{
      chip.className = 'active-filter-chip genre-chip';
      const node = nodeMap.get(f.label);
      chip.style.background = node ? node.color : '#666';
    }} else {{
      chip.className = 'active-filter-chip label-chip';
      chip.setAttribute('data-cat', f.category);
    }}
    chip.innerHTML = `${{f.label}} <span class="remove" data-idx="${{idx}}">&times;</span>`;
    chip.querySelector('.remove').addEventListener('click', (e) => {{
      e.stopPropagation();
      activeFilters.splice(idx, 1);
      onFiltersChanged();
    }});
    container.appendChild(chip);
  }});

  // Match count
  const count = document.createElement('span');
  count.className = 'filter-match-count';
  const matchCount = getFilteredMixes().length;
  count.textContent = `${{matchCount}} mix${{matchCount !== 1 ? 'es' : ''}}`;
  container.appendChild(count);

  // Clear all
  const clear = document.createElement('button');
  clear.className = 'clear-filters-btn';
  clear.textContent = 'Clear all';
  clear.addEventListener('click', clearAllFilters);
  container.appendChild(clear);
}}

function toggleLabelFilter(cat, label) {{
  const idx = activeFilters.findIndex(f => f.category === cat && f.label === label);
  if (idx >= 0) {{
    activeFilters.splice(idx, 1);
  }} else {{
    activeFilters.push({{ category: cat, label: label }});
  }}
  onFiltersChanged();
}}

function onFiltersChanged() {{
  renderActiveFilters();
  renderFilterPanel();
  applyLabelFilters();
}}

function applyLabelFilters() {{
  if (activeFilters.length === 0) {{
    // Collapse detail panel when no filters
    detailPanel.classList.add('collapsed');
    return;
  }}

  const filtered = getFilteredMixes();
  detailHistory = [{{ type: 'filters' }}];

  // Show filtered mixes in right detail panel (like genre detail)
  detailPanel.classList.remove('collapsed');
  document.getElementById('graphSearch')?.classList.add('hidden-by-panel');

  const filterDesc = activeFilters.map(f => f.label).join(' + ');
  const epList = filtered.slice(0, 50).map(m =>
    `<div class="track-item" style="cursor:pointer" onclick="selectEpisode(mixMap.get('${{m.id}}'))">` +
    `<span class="track-num">${{m.mix_number}}</span>` +
    `<span class="track-artist">${{m.artist}}</span></div>`
  ).join('');

  detailContent.innerHTML = `
    <div style="position:relative">
      <button class="detail-close" onclick="clearAllFilters()">Close</button>
      <div class="detail-header">
        <div class="ep-title">Label Filter</div>
        <div class="ep-artist">${{filterDesc}}</div>
        <div class="detail-meta">
          <div class="meta-item"><span class="meta-val">${{filtered.length}} episodes</span></div>
        </div>
      </div>
    </div>
    <div class="tracklist-section">
      <div class="tracklist-header">Episodes (${{filtered.length}})</div>
      ${{epList}}
      ${{filtered.length > 50 ? '<div class="no-tracklist">...and ' + (filtered.length - 50) + ' more</div>' : ''}}
    </div>
  `;

  detailPanel.scrollTop = 0;
}}

function clearAllFilters() {{
  activeFilters = [];
  expandedCats.clear();
  labelSearchQuery = '';
  if (labelSearchInput) labelSearchInput.value = '';
  onFiltersChanged();
}}

function showLabelMixes(cat, label) {{
  detailHistory = [{{ type: 'label', cat: cat, label: label }}];
  // Find all mixes with this label in this category
  const matches = MIXES.filter(m => {{
    const cats = m.label_categories || {{}};
    return (cats[cat] || []).includes(label);
  }});

  detailPanel.classList.remove('collapsed');
  document.getElementById('graphSearch')?.classList.add('hidden-by-panel');

  const epList = matches.slice(0, 50).map(m =>
    `<div class="track-item" style="cursor:pointer" onclick="selectEpisode(mixMap.get('${{m.id}}'))">` +
    `<span class="track-num">${{m.mix_number}}</span>` +
    `<span class="track-artist">${{m.artist}}</span></div>`
  ).join('');

  detailContent.innerHTML = `
    <div style="position:relative">
      <button class="detail-close" onclick="clearSelection()">Close</button>
      <div class="detail-header">
        <div class="ep-title"><span class="label-chip" data-cat="${{cat}}">${{label}}</span></div>
        <div class="detail-meta">
          <div class="meta-item"><span class="meta-val">${{matches.length}} episodes</span></div>
        </div>
      </div>
    </div>
    <div class="tracklist-section">
      <div class="tracklist-header">Episodes (${{matches.length}})</div>
      ${{epList}}
      ${{matches.length > 50 ? '<div class="no-tracklist">...and ' + (matches.length - 50) + ' more</div>' : ''}}
    </div>
  `;

  detailPanel.scrollTop = 0;
}}

// Tab switching (unified for desktop + mobile)
function switchTab(view) {{
  document.querySelectorAll('.center-tab').forEach(t => t.classList.toggle('active', t.dataset.view === view));
  document.querySelectorAll('.mobile-tab').forEach(t => t.classList.toggle('active', t.dataset.view === view));
  document.getElementById('graphArea').style.display = view === 'graph' ? '' : 'none';
  document.getElementById('labelFilterArea').style.display = view === 'labels' ? '' : 'none';
  const mixesArea = document.getElementById('mixesArea');
  if (mixesArea) mixesArea.classList.toggle('active-view', view === 'mixes');
  if (view === 'labels') renderFilterPanel();
  if (view === 'graph') setTimeout(resizeGraph, 100);
  if (view === 'mixes') renderMixesList();
}}

document.querySelectorAll('.center-tab').forEach(tab => {{
  tab.addEventListener('click', () => switchTab(tab.dataset.view));
}});

// Mobile tab bar switching
document.querySelectorAll('.mobile-tab').forEach(tab => {{
  tab.addEventListener('click', () => switchTab(tab.dataset.view));
}});

// ── Selection State ─────────────────────────────────────────────────────────
let highlightedNodes = new Set();
let directGenres = new Set();
let highlightedEdges = new Set();
let activeFilter = 'all';
let selectedMixId = null;
let detailHistory = [];  // navigation stack for detail panel

function selectEpisode(mix) {{
  // Toggle: if already selected, clear
  if (selectedMixId === mix.id) {{
    clearSelection();
    return;
  }}
  selectedMixId = mix.id;

  // Active state in list
  episodeList.querySelectorAll('.ep-item').forEach(el => {{
    el.classList.toggle('active', el.dataset.id === mix.id);
  }});

  // Highlight direct genres + their neighbors
  directGenres = new Set(mix.genres);
  highlightedNodes = new Set(mix.genres);
  // Add neighbors of direct genres
  mix.genres.forEach(g => {{
    const neighbors = adjacency.get(g);
    if (neighbors) neighbors.forEach(n => highlightedNodes.add(n));
  }});

  highlightedEdges = new Set();
  GRAPH_EDGES.forEach((e, i) => {{
    if (highlightedNodes.has(e.source) && highlightedNodes.has(e.target)) {{
      highlightedEdges.add(i);
    }}
  }});

  // Build detail panel
  showEpisodeDetail(mix);
  updateVisuals();
}}

function showEpisodeDetail(mix) {{
  if (detailHistory.length > 0) {{
    detailHistory.push({{ type: 'episode', id: mix.id }});
  }}
  detailPanel.classList.remove('collapsed');
  document.getElementById('graphSearch')?.classList.add('hidden-by-panel');

  const dur = mix.duration || '';
  const chips = mix.genres.map(g => {{
    const node = nodeMap.get(g);
    const color = node ? node.color : '#666';
    return `<span class="genre-chip" style="background:${{color}}" onclick="selectGenreNode('${{g}}')">${{g}}</span>`;
  }}).join('');

  // Keywords chips
  const kwHtml = mix.keywords ? mix.keywords.split(',').map(k => k.trim()).filter(Boolean)
    .map(k => `<span class="kw-chip">${{k}}</span>`).join('') : '';

  // Labels chips — use label_categories if available, else fall back to flat labels
  let labelsHtml = '';
  if (mix.label_categories && Object.keys(mix.label_categories).length > 0) {{
    const chips = [];
    // Render in category order, skip 'other'
    CAT_ORDER.forEach(cat => {{
      const items = mix.label_categories[cat];
      if (items && items.length) {{
        items.forEach(l => chips.push(`<span class="label-chip" data-cat="${{cat}}" style="cursor:pointer" onclick="showLabelMixes('${{cat}}','${{l.replace(/'/g, "\\\\'")}}')">${{l}}</span>`));
      }}
    }});
    labelsHtml = chips.join('');
  }} else if (mix.labels && mix.labels.length) {{
    labelsHtml = mix.labels.map(l => `<span class="label-chip">${{l}}</span>`).join('');
  }}

  // Links
  const links = [];
  if (mix.streamingUrl) links.push(`<a href="#" onclick="event.preventDefault(); playMix(mixMap.get('${{mix.id}}'))" class="detail-link">Listen</a>`);
  if (mix.url) links.push(`<a href="${{mix.url}}" target="_blank" rel="noopener noreferrer" class="detail-link">RA page</a>`);
  if (mix.artist_id) links.push(`<a href="https://ra.co/dj/${{mix.artist_id}}" target="_blank" rel="noopener noreferrer" class="detail-link">Artist</a>`);
  const linksHtml = links.length ? `<div class="detail-links">${{links.join('')}}</div>` : '';

  // Tracklist
  let tracklistHtml = '';
  if (mix.tracks && mix.tracks.length > 0) {{
    const tracks = mix.tracks.map((t, i) => {{
      const num = `<span class="track-num">${{i + 1}}</span>`;
      const artist = t.artist ? `<span class="track-artist">${{t.artist}}</span> ` : '';
      const title = t.title ? `<span class="track-title">${{artist ? '- ' : ''}}${{t.title}}</span>` : '';
      const label = t.label ? ` <span class="track-label">[${{t.label}}]</span>` : '';
      return `<div class="track-item">${{num}}${{artist}}${{title}}${{label}}</div>`;
    }}).join('');
    tracklistHtml = `
      <div class="tracklist-header">Tracklist (${{mix.tracks.length}})</div>
      ${{tracks}}`;
  }} else {{
    tracklistHtml = '<div class="no-tracklist">No tracklist available</div>';
  }}

  const hasHistory = detailHistory.length > 1;
  detailContent.innerHTML = `
    ${{mix.imageUrl ? `<div class="detail-cover"><img src="${{mix.imageUrl}}" alt="${{mix.artist}}"></div>` : ''}}
    <div style="position:relative">
      <button class="detail-close" onclick="${{hasHistory ? 'goBack()' : 'clearSelection()'}}">${{hasHistory ? '\u2190 Back' : 'Close'}}</button>
      <div class="detail-header">
        <div class="ep-title">RA.${{mix.mix_number.padStart(3,'0')}}</div>
        <div class="ep-artist">${{mix.artist}}</div>
        <div class="detail-meta">
          ${{mix.date ? `<div class="meta-item"><span class="meta-val">${{mix.date}}</span></div>` : ''}}
          ${{dur ? `<div class="meta-item"><span class="meta-val">${{dur}}</span></div>` : ''}}
          <div class="meta-item"><span class="meta-val">${{mix.tracks ? mix.tracks.length : 0}} tracks</span></div>
        </div>
        ${{linksHtml}}
        <div class="genre-chips">${{chips}}</div>
        ${{labelsHtml ? `<div class="label-chips">${{labelsHtml}}</div>` : ''}}
        ${{kwHtml ? `<div class="kw-chips">${{kwHtml}}</div>` : ''}}
      </div>
    </div>
    ${{mix.blurb ? `<div class="detail-blurb">${{mix.blurb}}</div>` : ''}}
    ${{mix.article ? `<div class="detail-section-header">About</div><div class="detail-article">${{mix.article}}</div>` : ''}}
    ${{mix.qa ? `<div class="detail-section-header">Q&A</div><div class="detail-article detail-qa">${{mix.qa}}</div>` : ''}}
    ${{tracklistHtml}}
  `;

  // Scroll to top + resize graph after panel opens
  detailPanel.scrollTop = 0;
  setTimeout(resizeGraph, 250);
}}

function showGenreDetail(nodeId, pushHistory = true) {{
  if (pushHistory) detailHistory = [{{ type: 'genre', id: nodeId }}];
  detailPanel.classList.remove('collapsed');
  document.getElementById('graphSearch')?.classList.add('hidden-by-panel');

  const n = nodeMap.get(nodeId);
  const count = n ? (n.count || 0) : 0;
  const family = n ? n.family : '';
  const color = n ? n.color : '#666';
  const desc = n && n.description ? n.description : '';

  // Find episodes with this genre
  const eps = MIXES.filter(m => m.genres.includes(nodeId));

  const epList = eps.slice(0, 50).map(m =>
    `<div class="track-item" style="cursor:pointer" onclick="selectEpisode(mixMap.get('${{m.id}}'))">` +
    `<span class="track-num">${{m.mix_number}}</span>` +
    `<span class="track-artist">${{m.artist}}</span></div>`
  ).join('');

  detailContent.innerHTML = `
    <div style="position:relative">
      <button class="detail-close" onclick="clearSelection()">Close</button>
      <div class="detail-header">
        <div class="ep-title" style="color:${{color}}">${{nodeId}}</div>
        <div class="ep-artist">${{family}}</div>
        <div class="detail-meta">
          <div class="meta-item"><span class="meta-val">${{count}} episodes</span></div>
        </div>
      </div>
      ${{desc ? `<div class="detail-blurb">${{desc}}</div>` : ''}}
    </div>
    <div class="tracklist-section">
      <div class="tracklist-header">Episodes (${{eps.length}})</div>
      ${{epList}}
      ${{eps.length > 50 ? '<div class="no-tracklist">...and ' + (eps.length - 50) + ' more</div>' : ''}}
    </div>
  `;

  setTimeout(resizeGraph, 250);
}}

function selectGenreNode(nodeId) {{
  directGenres = new Set([nodeId]);
  highlightedNodes = new Set([nodeId]);
  const neighbors = adjacency.get(nodeId);
  if (neighbors) neighbors.forEach(n => highlightedNodes.add(n));

  highlightedEdges = new Set();
  GRAPH_EDGES.forEach((e, i) => {{
    if (highlightedNodes.has(e.source) && highlightedNodes.has(e.target)) {{
      highlightedEdges.add(i);
    }}
  }});

  selectedMixId = null;

  // Show genre detail in right panel (don't filter left sidebar)

  showGenreDetail(nodeId);
  updateVisuals();
}}

function goBack() {{
  detailHistory.pop();  // remove current
  const prev = detailHistory[detailHistory.length - 1];
  if (!prev) {{ clearSelection(); return; }}
  selectedMixId = null;
  if (prev.type === 'genre') {{
    showGenreDetail(prev.id, false);
  }} else if (prev.type === 'label') {{
    showLabelMixes(prev.cat, prev.label);
  }} else if (prev.type === 'filters') {{
    applyLabelFilters();
  }} else {{
    clearSelection();
  }}
}}

function clearSelection() {{
  selectedMixId = null;
  detailHistory = [];
  highlightedNodes = new Set();
  directGenres = new Set();
  highlightedEdges = new Set();
  renderEpisodeList(searchInput.value);
  detailPanel.classList.add('collapsed');
  document.getElementById('graphSearch')?.classList.remove('hidden-by-panel');
  detailContent.innerHTML = '<div class="detail-placeholder">Select an episode to see details</div>';
  setTimeout(resizeGraph, 250);
  updateVisuals();
}}

function resetApp() {{
  // Clear all filters
  activeFilters = [];
  expandedCats.clear();
  labelSearchQuery = '';
  if (labelSearchInput) labelSearchInput.value = '';
  renderActiveFilters();
  renderFilterPanel();

  // Clear selection
  selectedMixId = null;
  highlightedNodes = new Set();
  directGenres = new Set();
  highlightedEdges = new Set();
  updateVisuals();

  // Reset episode list and search
  searchInput.value = '';
  renderEpisodeList();

  // Collapse detail panel
  detailPanel.classList.add('collapsed');
  document.getElementById('graphSearch')?.classList.remove('hidden-by-panel');
  detailContent.innerHTML = '<div class="detail-placeholder">Select an episode to see details</div>';

  // Switch to Network Graph tab
  switchTab('graph');

  setTimeout(resizeGraph, 250);
}}

// ── D3 Graph (co-occurrence) ─────────────────────────────────────────────
const svg = d3.select('#graph');
const graphArea = document.getElementById('graphArea');
let width = graphArea.clientWidth;
let height = graphArea.clientHeight;

svg.attr('viewBox', [0, 0, width, height]);

const defs = svg.append('defs');

// Helper: blend two hex colors
function blendColors(c1, c2) {{
  const r1 = parseInt(c1.slice(1,3),16), g1 = parseInt(c1.slice(3,5),16), b1 = parseInt(c1.slice(5,7),16);
  const r2 = parseInt(c2.slice(1,3),16), g2 = parseInt(c2.slice(3,5),16), b2 = parseInt(c2.slice(5,7),16);
  const r = Math.round((r1+r2)/2), g = Math.round((g1+g2)/2), b = Math.round((b1+b2)/2);
  return `rgb(${{r}},${{g}},${{b}})`;
}}

const gRoot = svg.append('g');

let currentZoomK = 0.9;
const zoom = d3.zoom()
  .scaleExtent([0.2, 5])
  .on('zoom', (event) => {{
    gRoot.attr('transform', event.transform);
    currentZoomK = event.transform.k;
    updateSemanticZoom();
  }});
svg.call(zoom);

// Family gravity centers — arranged in a ring (hoisted for resizeGraph)
const families = [...new Set(GRAPH_NODES.map(n => n.family))];
const familyColor = {{}};
families.forEach(f => {{ familyColor[f] = GRAPH_NODES.find(n => n.family === f)?.color || '#666'; }});

const familyCenter = {{}};

function resizeGraph() {{
  width = graphArea.clientWidth;
  height = graphArea.clientHeight;
  svg.attr('viewBox', [0, 0, width, height]);

  // Recenter simulation forces
  const newCx = width / 2, newCy = height / 2;
  simulation.force('center', d3.forceCenter(newCx, newCy).strength(0.03));
  const newRingR = Math.min(width, height) * 0.25;
  families.forEach((f, i) => {{
    const angle = (2 * Math.PI * i / families.length) - Math.PI / 2;
    familyCenter[f] = {{ x: newCx + newRingR * Math.cos(angle), y: newCy + newRingR * Math.sin(angle) }};
  }});
  simulation.alpha(0.15).restart();
  // Update family district label positions
  if (typeof familyLabels !== 'undefined') {{
    familyLabels
      .attr('x', f => familyLabelPos(f).x)
      .attr('y', f => familyLabelPos(f).y);
  }}
}}

// Prepare simulation data
const simNodes = GRAPH_NODES.map(n => ({{ ...n }}));
const simEdges = GRAPH_EDGES.map((e, i) => ({{ ...e, index: i }}));
const cx = width / 2, cy = height / 2;

// Initialize family gravity centers
const ringR = Math.min(width, height) * 0.25;
families.forEach((f, i) => {{
  const angle = (2 * Math.PI * i / families.length) - Math.PI / 2;
  familyCenter[f] = {{ x: cx + ringR * Math.cos(angle), y: cy + ringR * Math.sin(angle) }};
}});

// Simulation — musicological links + soft family gravity
const simulation = d3.forceSimulation(simNodes)
  .force('link', d3.forceLink(simEdges)
    .id(d => d.id)
    .distance(d => 40 + 100 * (1 - d.weight / maxWeight))
    .strength(d => 0.15 + 0.5 * (d.weight / maxWeight))
  )
  .force('charge', d3.forceManyBody()
    .strength(d => -30 - 100 * Math.sqrt(d.count / maxCount))
  )
  .force('center', d3.forceCenter(cx, cy).strength(0.03))
  .force('familyX', d3.forceX(d => {{
    const fc = familyCenter[d.family];
    return fc ? fc.x : cx;
  }}).strength(0.08))
  .force('familyY', d3.forceY(d => {{
    const fc = familyCenter[d.family];
    return fc ? fc.y : cy;
  }}).strength(0.08))
  .force('collide', d3.forceCollide().radius(d => nodeRadius(d.count) + 4))
  .alphaDecay(0.03);

// Family district labels — positioned at familyCenter, behind everything
// Family district labels — pushed outward past the node clusters
const familyLabelG = gRoot.append('g').attr('class', 'family-labels-layer');
function familyLabelPos(f) {{
  const fc = familyCenter[f];
  if (!fc) return {{ x: cx, y: cy }};
  // Push 60% further from center
  const dx = fc.x - cx, dy = fc.y - cy;
  return {{ x: fc.x + dx * 0.6, y: fc.y + dy * 0.6 }};
}}
const familyLabels = familyLabelG.selectAll('text')
  .data(families)
  .join('text')
  .attr('class', 'family-label')
  .attr('x', f => familyLabelPos(f).x)
  .attr('y', f => familyLabelPos(f).y)
  .attr('fill', f => familyColor[f])
  .attr('opacity', 0.2)
  .attr('font-size', 26)
  .text(f => f);

// Draw edges — color blended from source/target family, brightness scales with strength
const linkG = gRoot.append('g');
const links = linkG.selectAll('line')
  .data(simEdges)
  .join('line')
  .attr('stroke', d => {{
    const sc = (nodeMap.get(d.source.id || d.source) || {{}}).color || '#00e5ff';
    const tc = (nodeMap.get(d.target.id || d.target) || {{}}).color || '#00e5ff';
    return sc === tc ? sc : blendColors(sc, tc);
  }})
  .attr('stroke-width', d => 0.5 + 3.5 * (d.weight / maxWeight))
  .attr('stroke-opacity', d => 0.08 + 0.3 * (d.weight / maxWeight));

// Draw nodes
const nodeG = gRoot.append('g');
const nodeGroups = nodeG.selectAll('g')
  .data(simNodes)
  .join('g')
  .attr('cursor', 'pointer')
  .call(d3.drag()
    .filter(event => event.pointerType === 'mouse')
    .on('start', (event, d) => {{ if (!event.active) simulation.alphaTarget(0.3).restart(); d.fx = d.x; d.fy = d.y; }})
    .on('drag', (event, d) => {{ d.fx = event.x; d.fy = event.y; }})
    .on('end', (event, d) => {{ if (!event.active) simulation.alphaTarget(0); d.fx = null; d.fy = null; }})
  );

// Invisible touch hit area — ensures min 22px radius (44px target) on mobile
nodeGroups.append('circle').attr('class', 'node-hit')
  .attr('r', d => Math.max(22, nodeRadius(d.count) + 6))
  .attr('fill', 'transparent')
  .attr('stroke', 'none');

// Glow circle — intensity scales with node importance
nodeGroups.append('circle').attr('class', 'node-glow')
  .attr('r', d => nodeRadius(d.count) + 6 + 8 * Math.sqrt(d.count / maxCount))
  .attr('fill', d => d.color || '#666')
  .attr('opacity', d => 0.08 + 0.2 * Math.sqrt(d.count / maxCount));

// Idle pulse on top nodes
const idleThreshold = [...simNodes].sort((a,b) => b.count - a.count)[4]?.count || 50;
nodeGroups.each(function(d) {{
  if (d.count >= idleThreshold) d3.select(this).select('.node-glow').classed('idle-pulse', true);
}});

// Pulse ring
nodeGroups.append('circle').attr('class', 'pulse-ring')
  .attr('r', d => nodeRadius(d.count) + 4)
  .attr('fill', 'none').attr('stroke', '#00e5ff')
  .attr('stroke-opacity', 0).attr('stroke-width', 0);

// Main circle
nodeGroups.append('circle').attr('class', 'node-circle')
  .attr('r', d => nodeRadius(d.count))
  .attr('fill', d => d.color || '#666')
  .attr('opacity', 0.9);

// Labels
nodeGroups.append('text').attr('class', 'node-label')
  .attr('dy', d => nodeRadius(d.count) + 14)
  .attr('font-size', d => d.count > 50 ? 12 : d.count > 10 ? 10 : 8)
  .attr('font-weight', d => d.count > 50 ? 600 : 400)
  .attr('opacity', d => d.count >= 10 ? 0.8 : 0)
  .text(d => d.id);

// Tooltip
const tooltip = document.getElementById('tooltip');
function positionTooltip(event) {{
  const tip = document.getElementById('tooltip');
  const tt = tip.getBoundingClientRect();
  const vw = window.innerWidth;
  const vh = window.innerHeight;
  let x = event.clientX + 12;
  let y = event.clientY - 10;
  if (x + tt.width > vw - 8) x = event.clientX - tt.width - 12;
  if (x < 8) x = 8;
  if (y + tt.height > vh - 8) y = event.clientY - tt.height - 10;
  if (y < 8) y = 8;
  tip.style.left = x + 'px';
  tip.style.top = y + 'px';
}}
nodeGroups.on('mouseover', function(event, d) {{
  d3.select(this).select('.node-label').attr('opacity', 1);
  d3.select(this).select('.node-glow').attr('opacity', 0.4);
  // Brighten family district label
  familyLabels.attr('opacity', f => f === d.family ? 0.45 : 0.08);
  // Build tooltip with top connections
  const neighbors = adjacency.get(d.id);
  let conns = '';
  if (neighbors) {{
    const top = GRAPH_EDGES
      .filter(e => e.source.id === d.id || e.target.id === d.id)
      .sort((a, b) => b.weight - a.weight)
      .slice(0, 5)
      .map(e => {{
        const other = e.source.id === d.id ? e.target.id : e.source.id;
        return other;
      }});
    if (top.length) conns = ' → ' + top.join(', ');
  }}
  tooltip.classList.remove('has-desc');
  tooltip.textContent = `${{d.id}} — ${{d.count}} mixes${{conns}}`;
  tooltip.style.display = 'block';
  const ev = event;
  requestAnimationFrame(() => positionTooltip(ev));
}})
.on('mousemove', function(event) {{
  positionTooltip(event);
}})
.on('mouseout', function(event, d) {{
  if (!highlightedNodes.has(d.id) && d.count < 10) {{
    d3.select(this).select('.node-label').attr('opacity', 0);
  }}
  d3.select(this).select('.node-glow').attr('opacity', 0.08 + 0.2 * Math.sqrt(d.count / maxCount));
  familyLabels.attr('opacity', 0.2);
  tooltip.style.display = 'none';
}})
.on('click', function(event, d) {{
  event.stopPropagation();
  selectGenreNode(d.id);
}});

// Clear selection on background click — but NOT after pan/drag
let svgPointerDown = null;
svg.on('pointerdown', (event) => {{
  svgPointerDown = {{ x: event.clientX, y: event.clientY }};
}});
svg.on('click', (event) => {{
  if (svgPointerDown) {{
    const dx = Math.abs(event.clientX - svgPointerDown.x);
    const dy = Math.abs(event.clientY - svgPointerDown.y);
    if (dx > 5 || dy > 5) return; // was a drag, not a tap
  }}
  clearSelection();
}});

// ── Graph Search ──────────────────────────────────────────────
(function() {{
  const gsInput = document.getElementById('graphSearchInput');
  const gsClear = document.getElementById('graphSearchClear');
  const gsDropdown = document.getElementById('graphSearchDropdown');
  let gsActiveIdx = -1;

  function gsFilter(q) {{
    if (!q) return [];
    const lower = q.toLowerCase();
    return simNodes
      .filter(n => n.id.toLowerCase().includes(lower))
      .sort((a, b) => {{
        const aStarts = a.id.toLowerCase().startsWith(lower) ? 0 : 1;
        const bStarts = b.id.toLowerCase().startsWith(lower) ? 0 : 1;
        if (aStarts !== bStarts) return aStarts - bStarts;
        return b.count - a.count;
      }})
      .slice(0, 12);
  }}

  function gsRender(results) {{
    if (!results.length) {{ gsDropdown.style.display = 'none'; return; }}
    gsDropdown.innerHTML = results.map((n, i) =>
      `<div class="graph-search-item${{i === gsActiveIdx ? ' active' : ''}}" data-id="${{n.id}}">` +
      `<span class="gs-dot" style="background:${{n.color}}"></span>` +
      `<span>${{n.id}}</span>` +
      `<span class="gs-family">${{n.family}}</span>` +
      `<span class="gs-count">${{n.count}}</span>` +
      `</div>`
    ).join('');
    gsDropdown.style.display = 'block';
    gsDropdown.querySelectorAll('.graph-search-item').forEach(item => {{
      item.addEventListener('click', () => gsSelect(item.dataset.id));
    }});
  }}

  function gsSelect(nodeId) {{
    gsInput.value = '';
    gsClear.style.display = 'none';
    gsDropdown.style.display = 'none';
    gsActiveIdx = -1;
    gsInput.blur();

    // Zoom to node + select
    const node = simNodes.find(n => n.id === nodeId);
    if (node) {{
      const svgEl = document.getElementById('graph');
      const w = svgEl.clientWidth;
      const h = svgEl.clientHeight;
      const scale = 2.5;
      const tx = w / 2 - node.x * scale;
      const ty = h / 2 - node.y * scale;
      svg.transition().duration(600).call(
        zoom.transform,
        d3.zoomIdentity.translate(tx, ty).scale(scale)
      );
    }}
    setTimeout(() => selectGenreNode(nodeId), 150);
  }}

  gsInput.addEventListener('input', () => {{
    const q = gsInput.value.trim();
    gsClear.style.display = q ? 'block' : 'none';
    gsActiveIdx = -1;
    gsRender(gsFilter(q));
  }});

  gsInput.addEventListener('keydown', (e) => {{
    const items = gsDropdown.querySelectorAll('.graph-search-item');
    if (!items.length) return;
    if (e.key === 'ArrowDown') {{
      e.preventDefault();
      gsActiveIdx = Math.min(gsActiveIdx + 1, items.length - 1);
      gsRender(gsFilter(gsInput.value.trim()));
    }} else if (e.key === 'ArrowUp') {{
      e.preventDefault();
      gsActiveIdx = Math.max(gsActiveIdx - 1, 0);
      gsRender(gsFilter(gsInput.value.trim()));
    }} else if (e.key === 'Enter' && gsActiveIdx >= 0) {{
      e.preventDefault();
      gsSelect(items[gsActiveIdx].dataset.id);
    }} else if (e.key === 'Escape') {{
      gsDropdown.style.display = 'none';
      gsInput.blur();
    }}
  }});

  gsClear.addEventListener('click', () => {{
    gsInput.value = '';
    gsClear.style.display = 'none';
    gsDropdown.style.display = 'none';
    gsActiveIdx = -1;
    gsInput.focus();
  }});

  document.addEventListener('click', (e) => {{
    if (!e.target.closest('.graph-search')) gsDropdown.style.display = 'none';
  }});
}})();

simulation.on('tick', () => {{
  links.attr('x1', d => d.source.x).attr('y1', d => d.source.y)
    .attr('x2', d => d.target.x).attr('y2', d => d.target.y);
  nodeGroups.attr('transform', d => `translate(${{d.x}},${{d.y}})`);
}});

// ── Visual Update ───────────────────────────────────────────────────────────
function updateVisuals() {{
  const has = highlightedNodes.size > 0;

  nodeGroups.each(function(d) {{
    const el = d3.select(this);
    const isDirect = directGenres.has(d.id);
    const isNeighbor = highlightedNodes.has(d.id);

    const nodeOpacity = !has ? 0.9
      : isDirect ? 1
      : isNeighbor ? 0.4
      : 0.06;
    el.select('.node-circle').attr('opacity', nodeOpacity);
    const defaultGlow = 0.08 + 0.2 * Math.sqrt(d.count / maxCount);
    el.select('.node-glow')
      .attr('opacity', isDirect && has ? 0.5 : has ? 0.03 : defaultGlow);

    el.select('.pulse-ring')
      .classed('pulse', isDirect && has)
      .attr('stroke-opacity', isDirect && has ? 0.8 : 0);
    // Pause idle pulse during highlight
    el.select('.node-glow').classed('idle-pulse', !has && d.count >= idleThreshold);

    const labelOpacity = !has ? (d.count >= 10 ? 1 : 0)
      : isDirect ? 1
      : isNeighbor ? 0.4
      : 0;
    el.select('.node-label').attr('opacity', labelOpacity);
  }});

  // Family district labels
  const directFamily = has ? [...directGenres].map(id => nodeMap.get(id)?.family).filter(Boolean) : [];
  familyLabels.attr('opacity', f => {{
    if (!has) return 0.2;
    return directFamily.includes(f) ? 0.45 : 0.06;
  }});

  if (!has) {{
    updateSemanticZoom();
  }} else {{
    links.attr('stroke-opacity', (d, i) => {{
      return highlightedEdges.has(i) ? 0.3 + 0.5 * (d.weight / maxWeight) : 0.02;
    }});
  }}
}}

// ── Semantic Zoom ────────────────────────────────────────────────────────
function updateSemanticZoom() {{
  if (highlightedNodes.size > 0) return;

  const k = currentZoomK;
  const isMobile = window.innerWidth <= 768;

  // Labels: threshold drops as you zoom in (mobile-aware thresholds)
  let countThreshold;
  if (isMobile) {{
    countThreshold = k < 0.3 ? 999
      : k < 0.5 ? 100
      : k < 0.8 ? 30
      : k < 1.2 ? 8
      : k < 2 ? 3
      : 0;
  }} else {{
    countThreshold = k < 0.5 ? 999
      : k < 0.8 ? 50
      : k < 1.2 ? 10
      : k < 2 ? 3
      : 0;
  }}

  nodeGroups.each(function(d) {{
    const show = d.count >= countThreshold;
    d3.select(this).select('.node-label').attr('opacity', show ? 0.8 : 0);
  }});

  // Links: opacity grows with zoom
  const linkBoost = Math.max(0.3, Math.min(1.5, k / 1.5));
  links.attr('stroke-opacity', d => (0.08 + 0.25 * (d.weight / maxWeight)) * linkBoost);
}}

// ── Player (in detail panel) ─────────────────────────────────────────────
function getEmbedInfo(streamingUrl) {{
  if (!streamingUrl) return null;
  if (streamingUrl.includes('soundcloud.com')) {{
    const encoded = encodeURIComponent(streamingUrl);
    return {{
      type: 'sc',
      url: `https://w.soundcloud.com/player/?url=${{encoded}}&color=%23e63946&auto_play=true&hide_related=true&show_comments=false&show_user=false&show_reposts=false&show_teaser=false&visual=false`
    }};
  }}
  if (streamingUrl.includes('mixcloud.com')) {{
    const path = new URL(streamingUrl).pathname;
    return {{
      type: 'mc',
      url: `https://www.mixcloud.com/widget/iframe/?feed=${{encodeURIComponent(path)}}&autoplay=1&mini=0&hide_cover=1&light=1`
    }};
  }}
  return null;
}}

let currentlyPlayingMix = null;
function playMix(mix) {{
  currentlyPlayingMix = mix;
  const info = getEmbedInfo(mix.streamingUrl);
  const el = document.getElementById('sidebarPlayer');
  if (!info || !el) {{
    if (mix.streamingUrl) window.open(mix.streamingUrl, '_blank');
    return;
  }}
  el.className = `sidebar-player ${{info.type}}`;
  el.innerHTML = `<div class="player-info" style="cursor:pointer" onclick="if(currentlyPlayingMix) selectEpisode(currentlyPlayingMix)"><span class="player-ep">RA.${{mix.mix_number.padStart(3,'0')}}</span> <span class="player-artist">${{mix.artist}}</span></div>` +
    `<iframe src="${{info.url}}" allow="autoplay"></iframe>`;
}}

// ── Keyboard navigation ──────────────────────────────────────────────────
document.addEventListener('keydown', (e) => {{
  if (e.target.tagName === 'INPUT') return; // don't hijack search
  if (e.key !== 'ArrowUp' && e.key !== 'ArrowDown') return;
  if (!selectedMixId) return;

  e.preventDefault();
  const items = Array.from(episodeList.querySelectorAll('.ep-item'));
  if (!items.length) return;

  const currentIdx = items.findIndex(el => el.dataset.id === selectedMixId);
  let nextIdx;
  if (e.key === 'ArrowDown') {{
    nextIdx = currentIdx < items.length - 1 ? currentIdx + 1 : currentIdx;
  }} else {{
    nextIdx = currentIdx > 0 ? currentIdx - 1 : currentIdx;
  }}

  if (nextIdx !== currentIdx) {{
    const nextId = items[nextIdx].dataset.id;
    const nextMix = mixMap.get(nextId);
    if (nextMix) {{
      selectEpisode(nextMix);
      items[nextIdx].scrollIntoView({{ block: 'nearest' }});
    }}
  }}
}});


// ── Mobile: Mixes Tab ──────────────────────────────────────────────────────
function formatMixDate(dateStr) {{
  const d = new Date(dateStr);
  return d.toLocaleDateString('en-GB', {{ day: 'numeric', month: 'short' }});
}}

function renderMixesList() {{
  const list = document.getElementById('mixesList');
  const detail = document.getElementById('mixesDetail');
  if (!list) return;
  if (detail) detail.style.display = 'none';
  list.style.display = '';

  const sorted = [...MIXES].sort((a, b) => {{
    if (!a.date) return 1;
    if (!b.date) return -1;
    return b.date.localeCompare(a.date);
  }});

  list.innerHTML = sorted.map(m => {{
    const num = `RA.${{m.mix_number.padStart(3, '0')}}`;
    const date = m.date ? formatMixDate(m.date) : '';
    return `<div class="mix-row" data-id="${{m.id}}">
      <span class="mix-row-num">${{num}}</span>
      <span class="mix-row-artist">${{m.artist}}</span>
      <span class="mix-row-date">${{date}}</span>
    </div>`;
  }}).join('');

  list.querySelectorAll('.mix-row').forEach(row => {{
    row.addEventListener('click', () => {{
      const mix = mixMap.get(row.dataset.id);
      if (mix) showMixesDetail(mix);
    }});
  }});
}}

function showMixesDetail(mix) {{
  const mixesList = document.getElementById('mixesList');
  let mixesDetail = document.getElementById('mixesDetail');
  const mixesArea = document.getElementById('mixesArea');
  if (!mixesList || !mixesDetail || !mixesArea) return;

  mixesList.style.display = 'none';

  // Reset scroll on parent container
  mixesDetail.scrollTop = 0;

  // Create detail header + content
  const newHTML = `
    <div class="mixes-detail-header">
      <button class="mixes-back-btn" onclick="backToMixesList()">\u2190 Back</button>
    </div>
    <div class="mixes-detail-content"></div>
  `;

  mixesDetail.innerHTML = newHTML;

  // Get the content div
  const container = mixesDetail.querySelector('.mixes-detail-content');

  // Reset scroll again
  mixesDetail.scrollTop = 0;
  // Build the same detail content as showEpisodeDetail
  const dur = mix.duration || '';
  const chips = mix.genres.map(g => {{
    const node = nodeMap.get(g);
    const color = node ? node.color : '#666';
    return `<span class="genre-chip" style="background:${{color}}" onclick="selectGenreNode('${{g}}')">${{g}}</span>`;
  }}).join('');

  const kwHtml = mix.keywords ? mix.keywords.split(',').map(k => k.trim()).filter(Boolean)
    .map(k => `<span class="kw-chip">${{k}}</span>`).join('') : '';

  let labelsHtml = '';
  if (mix.label_categories && Object.keys(mix.label_categories).length > 0) {{
    const lchips = [];
    CAT_ORDER.forEach(cat => {{
      const items = mix.label_categories[cat];
      if (items && items.length) {{
        items.forEach(l => lchips.push(`<span class="label-chip" data-cat="${{cat}}" style="cursor:pointer" onclick="showLabelMixes('${{cat}}','${{l.replace(/'/g, "\\\\'")}}' )">${{l}}</span>`));
      }}
    }});
    labelsHtml = lchips.join('');
  }}

  const links = [];
  if (mix.streamingUrl) links.push(`<a href="${{mix.streamingUrl}}" target="_blank" rel="noopener" class="detail-link">Listen</a>`);
  if (mix.url) links.push(`<a href="${{mix.url}}" target="_blank" rel="noopener noreferrer" class="detail-link">RA page</a>`);
  if (mix.artist_id) links.push(`<a href="https://ra.co/dj/${{mix.artist_id}}" target="_blank" rel="noopener noreferrer" class="detail-link">Artist</a>`);
  const linksHtml = links.length ? `<div class="detail-links">${{links.join('')}}</div>` : '';

  let tracklistHtml = '';
  if (mix.tracks && mix.tracks.length > 0) {{
    const tracks = mix.tracks.map((t, i) => {{
      const num = `<span class="track-num">${{i + 1}}</span>`;
      const artist = t.artist ? `<span class="track-artist">${{t.artist}}</span> ` : '';
      const title = t.title ? `<span class="track-title">${{artist ? '- ' : ''}}${{t.title}}</span>` : '';
      const label = t.label ? ` <span class="track-label">[${{t.label}}]</span>` : '';
      return `<div class="track-item">${{num}}${{artist}}${{title}}${{label}}</div>`;
    }}).join('');
    tracklistHtml = `<div class="tracklist-header">Tracklist (${{mix.tracks.length}})</div>${{tracks}}`;
  }} else {{
    tracklistHtml = '<div class="no-tracklist">No tracklist available</div>';
  }}

  container.innerHTML = `
    ${{mix.imageUrl ? `<div class="detail-cover"><img src="${{mix.imageUrl}}" alt="${{mix.artist}}"></div>` : ''}}
    <div class="detail-header" style="padding: 16px">
      <div class="ep-title">RA.${{mix.mix_number.padStart(3,'0')}}</div>
      <div class="ep-artist">${{mix.artist}}</div>
      <div class="detail-meta">
        ${{mix.date ? `<div class="meta-item"><span class="meta-val">${{mix.date}}</span></div>` : ''}}
        ${{dur ? `<div class="meta-item"><span class="meta-val">${{dur}}</span></div>` : ''}}
        <div class="meta-item"><span class="meta-val">${{mix.tracks ? mix.tracks.length : 0}} tracks</span></div>
      </div>
      ${{linksHtml}}
      <div class="genre-chips">${{chips}}</div>
      ${{labelsHtml ? `<div class="label-chips">${{labelsHtml}}</div>` : ''}}
      ${{kwHtml ? `<div class="kw-chips">${{kwHtml}}</div>` : ''}}
    </div>
    ${{mix.blurb ? `<div class="detail-blurb">${{mix.blurb}}</div>` : ''}}
    ${{mix.article ? `<div class="detail-section-header">About</div><div class="detail-article">${{mix.article}}</div>` : ''}}
    ${{mix.qa ? `<div class="detail-section-header">Q&A</div><div class="detail-article detail-qa">${{mix.qa}}</div>` : ''}}
    ${{tracklistHtml}}
  `;

}}

function backToMixesList() {{
  const mixesList = document.getElementById('mixesList');
  const mixesDetail = document.getElementById('mixesDetail');
  if (mixesList) mixesList.style.display = '';
  if (mixesDetail) mixesDetail.style.display = 'none';
}}

// ── Mobile: Sheet handle swipe to expand/collapse ──────────────────────────
(function() {{
  const handle = detailPanel.querySelector('.sheet-handle');
  if (!handle) return;
  let startY = 0;
  handle.addEventListener('touchstart', (e) => {{
    startY = e.touches[0].clientY;
  }});
  handle.addEventListener('touchmove', (e) => {{
    e.preventDefault();
  }});
  handle.addEventListener('touchend', (e) => {{
    const endY = e.changedTouches[0].clientY;
    const dy = startY - endY;
    if (dy > 40) {{
      // Swipe up → expand
      detailPanel.classList.add('expanded');
    }} else if (dy < -40) {{
      // Swipe down → collapse or close
      if (detailPanel.classList.contains('expanded')) {{
        detailPanel.classList.remove('expanded');
      }} else {{
        clearSelection();
      }}
    }}
  }});
}})();

// ── Mobile: open streaming URL directly instead of embed ───────────────────
(function() {{
  const origPlayMix = playMix;
  playMix = function(mix) {{
    if (window.innerWidth <= 768) {{
      if (mix.streamingUrl) window.open(mix.streamingUrl, '_blank');
      return;
    }}
    origPlayMix(mix);
  }};
}})();
</script>
</body>
</html>"""

    return html


def main():
    print("Loading mixes from Excel...")
    mixes = load_mixes()
    print(f"  {len(mixes)} episodes loaded")

    print("Loading genre musicology graph...")
    with open(DATA_DIR / "genre_musicology.json") as f:
        graph = json.load(f)
    print(f"  {len(graph['nodes'])} nodes, {len(graph['edges'])} edges")

    check_gaps(mixes, graph)

    print("Building HTML...")
    html = build_html(mixes, graph)

    out_path = ROOT / "ra_genre_network.html"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  → {out_path} ({len(html)//1024} KB)")


if __name__ == "__main__":
    main()
