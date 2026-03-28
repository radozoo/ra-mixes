"""
excel_exporter.py — exportuje všetky dáta do data/ra_full_export.xlsx.

Sheets:
  Episodes      — základné info o každej epizóde
  Tracks        — tracklist záznamy
  Genre Edges   — čisté (normalizované) genre hrany
  Genre Summary — počet epizód na žáner
  Summary       — celkové štatistiky
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent.parent / "data"
OUTPUT_PATH = DATA_DIR / "ra_full_export.xlsx"


def _auto_width(ws, max_width: int = 60):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _header_row(ws, headers: list[str]):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEEFF")
        cell.alignment = Alignment(horizontal="center")


def load_jsonl(path: Path) -> list[dict]:
    rows = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
    return rows


def export(output_path: Path = OUTPUT_PATH):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # --- Episodes ---
    episodes = load_jsonl(DATA_DIR / "episodes.jsonl")
    ws_ep = wb.create_sheet("Episodes")
    ep_headers = ["podcast_id", "title", "artist_name", "date", "duration_seconds",
                  "has_tracklist", "scrape_quality", "scrape_source", "blurb"]
    _header_row(ws_ep, ep_headers)
    for ep in sorted(episodes, key=lambda x: int(x.get("podcast_id", 0) or 0)):
        ws_ep.append([ep.get(h) for h in ep_headers])
    _auto_width(ws_ep)

    # --- Tracks ---
    tracks = load_jsonl(DATA_DIR / "tracks.jsonl")
    ws_tr = wb.create_sheet("Tracks")
    tr_headers = ["track_id", "podcast_id", "position", "artist", "title", "label",
                  "release_year", "remix"]
    _header_row(ws_tr, tr_headers)
    for t in sorted(tracks, key=lambda x: (
        int(x.get("podcast_id", 0) or 0),
        int(x.get("position", 0) or 0),
    )):
        ws_tr.append([t.get(h) for h in tr_headers])
    _auto_width(ws_tr)

    # --- Genre Edges (clean) ---
    clean_path = DATA_DIR / "genre_edges_clean.jsonl"
    genre_edges = load_jsonl(clean_path if clean_path.exists() else DATA_DIR / "genre_edges.jsonl")
    ws_ge = wb.create_sheet("Genre Edges")
    ge_headers = ["entity_id", "genre_canonical", "genre_raw", "source", "confidence"]
    _header_row(ws_ge, ge_headers)
    for e in sorted(genre_edges, key=lambda x: int(x.get("entity_id", 0) or 0)):
        ws_ge.append([e.get(h) for h in ge_headers])
    _auto_width(ws_ge)

    # --- Genre Summary ---
    from collections import Counter
    genre_counts: Counter = Counter()
    for e in genre_edges:
        genre_counts[e.get("genre_canonical", "")] += 1
    ws_gs = wb.create_sheet("Genre Summary")
    _header_row(ws_gs, ["genre", "episode_count"])
    for g, c in genre_counts.most_common():
        ws_gs.append([g, c])
    _auto_width(ws_gs)

    # --- Summary ---
    ws_sum = wb.create_sheet("Summary")
    _header_row(ws_sum, ["metric", "value"])
    ep_with_tracklist = sum(1 for e in episodes if e.get("has_tracklist"))
    summary = [
        ("Episodes total", len(episodes)),
        ("Episodes with tracklist", ep_with_tracklist),
        ("Tracks total", len(tracks)),
        ("Genre edges (clean)", len(genre_edges)),
        ("Unique genres", len(genre_counts)),
        ("Artists (unique)", len(set(e.get("artist_name") for e in episodes if e.get("artist_name")))),
    ]
    for k, v in summary:
        ws_sum.append([k, v])
    _auto_width(ws_sum)

    wb.save(output_path)
    print(f"Exportované: {output_path}")
    print(f"  {len(episodes)} epizód | {len(tracks)} trackov | {len(genre_edges)} genre edges")
    return output_path


if __name__ == "__main__":
    export()
